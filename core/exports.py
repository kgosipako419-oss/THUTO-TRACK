"""School-wide data exports for Ministry submissions / archival.

Builders return ``(bytes, content_type, filename)`` tuples so views can pass
the bytes straight to an ``HttpResponse``.

Four exports + one combined "Ministry pack" workbook:

    marks       — long-format row-per-mark
    attendance  — long-format row-per-attendance-record
    roster      — one row per student enrolled in the year
    performance — per (class, subject, term) average

The combined pack puts all four sheets in a single styled .xlsx.

These layouts are designed to be friendly for re-mapping into a specific
Ministry template (e.g. BETP / SmartBots) — they are NOT a Ministry-defined
format. Adjust the column lists below if your Ministry requires a specific
shape.
"""

import csv
from io import BytesIO, StringIO

from django.db.models import Avg, Count, F, FloatField
from django.db.models.functions import Cast
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import Attendance, Mark, Student


HEADER_FILL = PatternFill("solid", fgColor="047857")
HEADER_FONT = Font(bold=True, color="FFFFFF")

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------------------
# Row builders (pure data — independent of output format)
# ---------------------------------------------------------------------------

def _marks_rows(school, year: int, term=None):
    qs = (
        Mark.objects.filter(student__school=school, academic_year=year)
        .select_related("student", "student__class_group", "subject", "teacher__user")
        .order_by(
            "student__class_group__grade_level",
            "student__class_group__name",
            "student__last_name",
            "student__first_name",
            "subject__name",
            "recorded_at",
        )
    )
    if term is not None:
        qs = qs.filter(term=term)
    headers = [
        "student_number", "first_name", "last_name",
        "class", "grade_level",
        "subject_code", "subject_name",
        "assessment_type", "title",
        "score", "max_score", "percentage",
        "term", "academic_year",
        "teacher", "recorded_at",
    ]
    rows = []
    for m in qs:
        rows.append([
            m.student.student_number,
            m.student.first_name,
            m.student.last_name,
            m.student.class_group.name,
            m.student.class_group.grade_level,
            m.subject.code,
            m.subject.name,
            m.get_assessment_type_display(),
            m.title,
            float(m.score),
            float(m.max_score),
            round(m.percentage, 1),
            m.term,
            m.academic_year,
            m.teacher.user.get_full_name() or m.teacher.user.username,
            m.recorded_at.strftime("%Y-%m-%d %H:%M"),
        ])
    return headers, rows


def _attendance_rows(school, year: int):
    qs = (
        Attendance.objects.filter(student__school=school, date__year=year)
        .select_related("student", "student__class_group", "recorded_by__user")
        .order_by(
            "student__class_group__grade_level",
            "student__class_group__name",
            "student__last_name",
            "student__first_name",
            "date",
        )
    )
    headers = [
        "student_number", "first_name", "last_name",
        "class", "grade_level",
        "date", "status", "recorded_by", "notes",
    ]
    rows = []
    for a in qs:
        rows.append([
            a.student.student_number,
            a.student.first_name,
            a.student.last_name,
            a.student.class_group.name,
            a.student.class_group.grade_level,
            a.date.isoformat(),
            a.get_status_display(),
            a.recorded_by.user.get_full_name() or a.recorded_by.user.username,
            a.notes,
        ])
    return headers, rows


def _roster_rows(school, year: int):
    qs = (
        Student.objects.filter(school=school, class_group__academic_year=year)
        .select_related("class_group")
        .order_by(
            "class_group__grade_level",
            "class_group__name",
            "last_name",
            "first_name",
        )
    )
    headers = [
        "student_number", "first_name", "last_name",
        "gender", "date_of_birth",
        "class", "grade_level", "academic_year",
        "parent_name", "parent_phone",
        "is_active", "enrolled_at",
    ]
    rows = []
    for s in qs:
        rows.append([
            s.student_number, s.first_name, s.last_name,
            s.get_gender_display(),
            s.date_of_birth.isoformat() if s.date_of_birth else "",
            s.class_group.name, s.class_group.grade_level, s.class_group.academic_year,
            s.parent_name, s.parent_phone,
            "Yes" if s.is_active else "No",
            s.enrolled_at.isoformat(),
        ])
    return headers, rows


def _performance_rows(school, year: int):
    qs = (
        Mark.objects.filter(student__school=school, academic_year=year)
        .annotate(
            pct=Cast(F("score"), FloatField()) * 100.0 / Cast(F("max_score"), FloatField()),
        )
        .values(
            "student__class_group__name",
            "student__class_group__grade_level",
            "subject__code", "subject__name", "term",
        )
        .annotate(
            avg_pct=Avg("pct"),
            assessment_count=Count("id"),
            student_count=Count("student", distinct=True),
        )
        .order_by(
            "student__class_group__grade_level",
            "student__class_group__name",
            "subject__name", "term",
        )
    )
    headers = [
        "class", "grade_level",
        "subject_code", "subject_name", "term",
        "assessment_count", "student_count", "average_percentage",
    ]
    rows = []
    for r in qs:
        rows.append([
            r["student__class_group__name"],
            r["student__class_group__grade_level"],
            r["subject__code"],
            r["subject__name"],
            r["term"],
            r["assessment_count"],
            r["student_count"],
            round(r["avg_pct"], 1) if r["avg_pct"] is not None else 0,
        ])
    return headers, rows


# ---------------------------------------------------------------------------
# Output formats
# ---------------------------------------------------------------------------

def _style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _autosize_columns(ws, num_cols):
    for col in range(1, num_cols + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in ws[letter]
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)


def _single_sheet_xlsx(headers, rows, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    _style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"
    _autosize_columns(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_bytes(headers, rows):
    out = StringIO()
    w = csv.writer(out)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    # BOM so Excel opens UTF-8 cleanly with diacritics
    return ("﻿" + out.getvalue()).encode("utf-8")


KIND_LABELS = {
    "marks": "Marks",
    "attendance": "Attendance",
    "roster": "Roster",
    "performance": "Performance",
}

VALID_KINDS = set(KIND_LABELS)
VALID_FORMATS = {"xlsx", "csv"}


def _rows_for(kind, school, year, term=None):
    if kind == "marks":
        return _marks_rows(school, year, term)
    if kind == "attendance":
        return _attendance_rows(school, year)
    if kind == "roster":
        return _roster_rows(school, year)
    if kind == "performance":
        return _performance_rows(school, year)
    raise ValueError(f"Unknown export kind: {kind}")


def build_single_export(kind: str, school, year: int, term=None, fmt: str = "xlsx"):
    """Single-kind export. Returns (bytes, content_type, filename)."""
    if kind not in VALID_KINDS:
        raise ValueError(f"Unknown export kind: {kind}")
    if fmt not in VALID_FORMATS:
        raise ValueError(f"Unknown export format: {fmt}")

    headers, rows = _rows_for(kind, school, year, term)
    safe_school = "".join(c if c.isalnum() else "_" for c in school.name).strip("_")[:40]
    suffix = f"_term{term}" if term is not None else ""

    if fmt == "csv":
        return (
            _csv_bytes(headers, rows),
            "text/csv; charset=utf-8",
            f"{safe_school}_{kind}_{year}{suffix}.csv",
        )
    return (
        _single_sheet_xlsx(headers, rows, f"{KIND_LABELS[kind]} {year}"),
        XLSX_CONTENT_TYPE,
        f"{safe_school}_{kind}_{year}{suffix}.xlsx",
    )


def build_ministry_pack(school, year: int, term=None):
    """Combined .xlsx with one sheet per export kind. Returns (bytes, content_type, filename)."""
    wb = Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("Summary")
    cover["A1"] = "ThutoTrack Ministry submission pack"
    cover["A1"].font = Font(bold=True, size=14)
    cover["A2"] = f"School: {school.name}"
    cover["A3"] = f"School code: {school.code}"
    cover["A4"] = f"Academic year: {year}"
    if term is not None:
        cover["A5"] = f"Term filter: Term {term} (marks only)"
    cover["A6"] = (
        "This workbook contains four sheets: Roster, Marks, Attendance and "
        "Performance. Adjust column mapping to match the Ministry template."
    )
    cover.column_dimensions["A"].width = 80

    # Order: roster first (so reader sees the population), then marks, attendance, performance
    for kind in ("roster", "marks", "attendance", "performance"):
        headers, rows = _rows_for(kind, school, year, term if kind == "marks" else None)
        ws = wb.create_sheet(KIND_LABELS[kind])
        ws.append(headers)
        for r in rows:
            ws.append(r)
        _style_header_row(ws, len(headers))
        ws.freeze_panes = "A2"
        _autosize_columns(ws, len(headers))

    buf = BytesIO()
    wb.save(buf)
    safe_school = "".join(c if c.isalnum() else "_" for c in school.name).strip("_")[:40]
    suffix = f"_term{term}" if term is not None else ""
    return (
        buf.getvalue(),
        XLSX_CONTENT_TYPE,
        f"{safe_school}_ministry_pack_{year}{suffix}.xlsx",
    )
