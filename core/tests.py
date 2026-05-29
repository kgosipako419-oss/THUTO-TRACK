"""Unit tests for core models, exports, and WhatsApp handler logic."""

from datetime import date
from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from core.exports import build_ministry_pack, build_single_export
from core.factories import (
    make_class,
    make_parent_account,
    make_school,
    make_student,
    make_subject,
    make_teacher,
    make_term_schedule,
)
from core.models import Attendance, BehaviorNote, Mark, ParentSession, Student
from core.whatsapp import (
    find_students_for_phone,
    handle_inbound,
    normalize_phone_for_username,
    phones_match,
)


class PhoneNormalizationTests(TestCase):
    def test_normalize_strips_formatting(self):
        for raw in ["+267 71 222 001", "26771222001", "071222001", "71222001"]:
            self.assertEqual(normalize_phone_for_username(raw), "26771222001")

    def test_normalize_empty(self):
        self.assertEqual(normalize_phone_for_username(""), "")
        self.assertEqual(normalize_phone_for_username(None), "")

    def test_phones_match_loose(self):
        self.assertTrue(phones_match("+267 71 222 001", "whatsapp:+26771222001"))
        self.assertTrue(phones_match("071222001", "+267 71 222 001"))
        self.assertFalse(phones_match("+267 71 222 001", "+267 71 222 002"))
        self.assertFalse(phones_match("", "+267 71 222 001"))


class StudentLookupTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.school = make_school()
        cls.cg = make_class(cls.school)
        cls.s1 = make_student(
            cls.school, cls.cg, student_number="S1",
            first_name="A", last_name="One",
            parent_phone="+267 71 222 001",
        )
        cls.s2 = make_student(
            cls.school, cls.cg, student_number="S2",
            first_name="B", last_name="Two",
            parent_phone="+267 71 222 001",
        )
        cls.s3 = make_student(
            cls.school, cls.cg, student_number="S3",
            first_name="C", last_name="Three",
            parent_phone="+267 71 999 999",
        )

    def test_finds_both_students_for_same_parent(self):
        kids = find_students_for_phone("whatsapp:+26771222001")
        self.assertEqual({s.id for s in kids}, {self.s1.id, self.s2.id})

    def test_finds_no_students_for_unknown_phone(self):
        self.assertEqual(find_students_for_phone("+26700000000"), [])

    def test_ignores_inactive_students(self):
        self.s1.is_active = False
        self.s1.save()
        kids = find_students_for_phone("+26771222001")
        self.assertEqual([s.id for s in kids], [self.s2.id])


class WhatsAppHandlerTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.school = make_school()
        cls.teacher = make_teacher(cls.school)
        cls.cg = make_class(cls.school, class_teacher=cls.teacher)
        cls.math = make_subject(cls.school, code="MATH", name="Mathematics")
        cls.english = make_subject(cls.school, code="ENG", name="English")
        cls.student = make_student(
            cls.school, cls.cg, student_number="WA-S1",
            first_name="Naledi", last_name="Tau",
            parent_phone="+267 71 222 001",
        )
        # Term schedule covering today's year so the report logic works
        year = date.today().year
        make_term_schedule(cls.school, term=1, year=year,
                           start=date(year, 1, 1), end=date(year, 12, 31))
        Mark.objects.create(
            student=cls.student, subject=cls.math, teacher=cls.teacher,
            assessment_type="TEST", title="Mid-term", score=80, max_score=100,
            term=1, academic_year=year,
        )
        Mark.objects.create(
            student=cls.student, subject=cls.english, teacher=cls.teacher,
            assessment_type="ASSIGN", title="Essay 1", score=60, max_score=100,
            term=1, academic_year=year,
        )
        Attendance.objects.create(
            student=cls.student, date=date(year, 2, 1), status="P",
            recorded_by=cls.teacher,
        )
        Attendance.objects.create(
            student=cls.student, date=date(year, 2, 2), status="A",
            recorded_by=cls.teacher,
        )
        BehaviorNote.objects.create(
            student=cls.student, teacher=cls.teacher,
            category="POS", note="Helped a classmate.",
        )

    def test_unknown_phone_gets_onboarding(self):
        reply = handle_inbound("+26799999999", "hi")
        self.assertIn("isn't linked to a student", reply)
        # No session is created for unknown numbers
        self.assertFalse(ParentSession.objects.filter(phone__contains="9999").exists())

    def test_greeting_shows_single_student_menu(self):
        reply = handle_inbound("+26771222001", "hi")
        self.assertIn("ThutoTrack", reply)
        self.assertIn("Naledi", reply)
        self.assertIn("marks", reply)

    def test_marks_returns_marks(self):
        reply = handle_inbound("+26771222001", "marks")
        self.assertIn("Mathematics", reply)
        self.assertIn("English", reply)
        self.assertIn("Mid-term", reply)

    def test_attendance_returns_correct_rate(self):
        reply = handle_inbound("+26771222001", "attendance")
        self.assertIn("Present: 1/2", reply)
        self.assertIn("50%", reply)

    def test_report_returns_overall_average(self):
        reply = handle_inbound("+26771222001", "report")
        self.assertIn("Overall", reply)
        self.assertIn("70%", reply)  # (80 + 60) / 2

    def test_behavior_returns_note(self):
        reply = handle_inbound("+26771222001", "behavior")
        self.assertIn("Helped a classmate", reply)

    def test_digit_shortcut_for_marks(self):
        reply = handle_inbound("+26771222001", "1")
        self.assertIn("Mathematics", reply)

    def test_unknown_command_falls_back_to_menu(self):
        reply = handle_inbound("+26771222001", "gibberish")
        self.assertIn("didn't understand", reply)
        self.assertIn("marks", reply)

    def test_session_auto_selects_single_student(self):
        handle_inbound("+26771222001", "hi")
        s = ParentSession.objects.get(phone="+26771222001")
        self.assertEqual(s.selected_student_id, self.student.id)

    def test_multi_student_requires_selection(self):
        # Add a sibling
        make_student(
            self.school, self.cg, student_number="WA-S2",
            first_name="Karabo", last_name="Tau",
            parent_phone="+267 71 222 001",
        )
        ParentSession.objects.filter(phone="+26771222001").delete()
        reply = handle_inbound("+26771222001", "hi")
        self.assertIn("Karabo", reply)
        self.assertIn("Naledi", reply)
        self.assertIn("number", reply.lower())

    def test_invalid_multi_student_choice_gentle_error(self):
        make_student(
            self.school, self.cg, student_number="WA-S2",
            first_name="Karabo", last_name="Tau",
            parent_phone="+267 71 222 001",
        )
        ParentSession.objects.filter(phone="+26771222001").delete()
        handle_inbound("+26771222001", "hi")  # show multi-student menu
        reply = handle_inbound("+26771222001", "99")
        self.assertIn("not one of your students", reply)


class ExportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.school = make_school()
        cls.teacher = make_teacher(cls.school)
        cls.cg = make_class(cls.school, class_teacher=cls.teacher)
        cls.math = make_subject(cls.school, code="MATH", name="Mathematics")
        cls.students = [
            make_student(cls.school, cls.cg, student_number=f"EX-{i}",
                         first_name=f"Stu{i}", last_name="Test")
            for i in range(3)
        ]
        for s in cls.students:
            Mark.objects.create(
                student=s, subject=cls.math, teacher=cls.teacher,
                assessment_type="TEST", title="Test 1",
                score=70, max_score=100, term=1, academic_year=2026,
            )
        Attendance.objects.create(
            student=cls.students[0], date=date(2026, 3, 1), status="P",
            recorded_by=cls.teacher,
        )

    def test_marks_xlsx(self):
        body, ct, name = build_single_export("marks", self.school, 2026, fmt="xlsx")
        self.assertTrue(name.endswith(".xlsx"))
        self.assertIn("openxmlformats", ct)
        wb = load_workbook(BytesIO(body))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        self.assertIn("student_number", headers)
        self.assertIn("percentage", headers)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 3)

    def test_marks_csv_has_bom(self):
        body, ct, name = build_single_export("marks", self.school, 2026, fmt="csv")
        self.assertTrue(name.endswith(".csv"))
        self.assertTrue(body.startswith("﻿".encode("utf-8")))
        self.assertIn(b"student_number", body)

    def test_attendance_export_counts(self):
        body, _, _ = build_single_export("attendance", self.school, 2026, fmt="xlsx")
        wb = load_workbook(BytesIO(body))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)

    def test_roster_export_includes_all_students(self):
        body, _, _ = build_single_export("roster", self.school, 2026, fmt="xlsx")
        wb = load_workbook(BytesIO(body))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 3)

    def test_performance_aggregates(self):
        body, _, _ = build_single_export("performance", self.school, 2026, fmt="xlsx")
        wb = load_workbook(BytesIO(body))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        # One (class, subject, term) bucket
        self.assertEqual(len(rows), 1)
        # average_percentage column is the last one
        self.assertEqual(rows[0][-1], 70.0)

    def test_term_filter_excludes_other_terms(self):
        body, _, _ = build_single_export("marks", self.school, 2026, term=2, fmt="xlsx")
        wb = load_workbook(BytesIO(body))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(rows, [])

    def test_pack_has_all_sheets(self):
        body, ct, name = build_ministry_pack(self.school, 2026)
        wb = load_workbook(BytesIO(body))
        self.assertSetEqual(
            set(wb.sheetnames),
            {"Summary", "Roster", "Marks", "Attendance", "Performance"},
        )

    def test_pack_cross_school_isolation(self):
        other = make_school(name="Other School", code="OTHER-T")
        other_cg = make_class(other)
        other_stu = make_student(other, other_cg, student_number="OTH-1",
                                 first_name="Off", last_name="School")
        other_sub = make_subject(other, code="MATH", name="Math")
        other_teacher = make_teacher(other, username="other_teach_e")
        Mark.objects.create(
            student=other_stu, subject=other_sub, teacher=other_teacher,
            assessment_type="TEST", title="Other test",
            score=99, max_score=100, term=1, academic_year=2026,
        )
        body, _, _ = build_single_export("marks", self.school, 2026, fmt="xlsx")
        wb = load_workbook(BytesIO(body))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        titles = {r[8] for r in rows}
        self.assertNotIn("Other test", titles)


class HealthCheckTests(TestCase):
    def test_healthz_returns_ok(self):
        resp = self.client.get("/healthz/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.content, b"ok")


class ParentAccountTests(TestCase):
    def test_make_parent_account_creates_user(self):
        user = make_parent_account("+267 71 222 001", "1234")
        self.assertEqual(user.username, "26771222001")
        self.assertTrue(user.check_password("1234"))
        self.assertEqual(user.role, "PARENT")

    def test_make_parent_account_idempotent(self):
        u1 = make_parent_account("+267 71 222 001", "1234")
        u2 = make_parent_account("+267 71 222 001", "5678")
        self.assertEqual(u1.pk, u2.pk)
        u1.refresh_from_db()
        self.assertTrue(u1.check_password("5678"))
