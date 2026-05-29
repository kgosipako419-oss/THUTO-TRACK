"""End-to-end tests for the school admin portal."""

from datetime import date
from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from core.factories import (
    make_class,
    make_school,
    make_school_admin,
    make_student,
    make_subject,
    make_teacher,
)
from core.models import (
    ClassGroup,
    Enquiry,
    Mark,
    Student,
    Subject,
    TeacherProfile,
    TermSchedule,
    User,
)


class AdminPortalBase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.school = make_school()
        cls.admin = make_school_admin(cls.school, username="head", password="headpass1")
        cls.teacher = make_teacher(cls.school, username="kgosi", password="teach1234")
        cls.subjects = [
            make_subject(cls.school, code="MATH", name="Mathematics"),
            make_subject(cls.school, code="ENG", name="English"),
        ]
        cls.teacher.subjects.set(cls.subjects)
        cls.cg = make_class(cls.school, class_teacher=cls.teacher)
        cls.teacher.classes_taught.add(cls.cg)
        cls.students = [
            make_student(cls.school, cls.cg, student_number=f"S-{i:03d}",
                         first_name=f"Stu{i}", last_name="Test",
                         parent_phone=f"+267 71 222 00{i}")
            for i in range(1, 4)
        ]

    def login(self):
        self.assertTrue(self.client.login(username="head", password="headpass1"))


class AuthAndAccessTests(AdminPortalBase):
    def test_login_required(self):
        resp = self.client.get("/admin-portal/")
        self.assertEqual(resp.status_code, 302)

    def test_dashboard_renders_for_admin(self):
        self.login()
        resp = self.client.get("/admin-portal/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.school.name)
        self.assertContains(resp, "School-wide performance")

    def test_teacher_denied_admin_portal(self):
        self.assertTrue(self.client.login(username="kgosi", password="teach1234"))
        resp = self.client.get("/admin-portal/")
        self.assertEqual(resp.status_code, 403)


class EnquiryInboxTests(AdminPortalBase):
    def test_inbox_lists_and_filters(self):
        self.login()
        Enquiry.objects.create(school=self.school, from_teacher=self.teacher,
                               category="HR", subject="Leave", body="Need leave")
        Enquiry.objects.create(school=self.school, from_teacher=self.teacher,
                               category="TECH", subject="Bug", body="App slow", status="PROG")
        resp = self.client.get("/admin-portal/enquiries/")
        self.assertContains(resp, "Leave")
        self.assertContains(resp, "Bug")
        resp = self.client.get("/admin-portal/enquiries/?status=OPEN")
        self.assertContains(resp, "Leave")
        self.assertNotContains(resp, "Bug")
        resp = self.client.get("/admin-portal/enquiries/?category=TECH")
        self.assertContains(resp, "Bug")
        self.assertNotContains(resp, "Leave")

    def test_respond_to_enquiry_sets_resolved_at(self):
        self.login()
        e = Enquiry.objects.create(school=self.school, from_teacher=self.teacher,
                                   category="HR", subject="Sub", body="b")
        self.client.post(
            f"/admin-portal/enquiries/{e.id}/",
            {"status": "DONE", "response": "Approved."},
        )
        e.refresh_from_db()
        self.assertEqual(e.status, "DONE")
        self.assertEqual(e.response, "Approved.")
        self.assertIsNotNone(e.resolved_at)

    def test_reopening_clears_resolved_at(self):
        self.login()
        e = Enquiry.objects.create(school=self.school, from_teacher=self.teacher,
                                   category="HR", subject="s", body="b")
        self.client.post(f"/admin-portal/enquiries/{e.id}/",
                         {"status": "DONE", "response": "r"})
        self.client.post(f"/admin-portal/enquiries/{e.id}/",
                         {"status": "PROG", "response": "r"})
        e.refresh_from_db()
        self.assertIsNone(e.resolved_at)

    def test_cross_school_enquiry_404(self):
        self.login()
        other = make_school(name="Other", code="OTH-1")
        other_t = make_teacher(other, username="other_t")
        e = Enquiry.objects.create(school=other, from_teacher=other_t,
                                   category="HR", subject="x", body="y")
        resp = self.client.get(f"/admin-portal/enquiries/{e.id}/")
        self.assertEqual(resp.status_code, 404)


class TeacherManagementTests(AdminPortalBase):
    def test_create_teacher_assigns_subjects_and_classes(self):
        self.login()
        resp = self.client.post(
            "/admin-portal/teachers/new/",
            {
                "first_name": "New", "last_name": "Teach",
                "username": "newt", "password": "securepw1",
                "email": "n@example.com", "employee_id": "T-99",
                "subjects": [str(self.subjects[0].id)],
                "classes": [str(self.cg.id)],
            }, follow=True,
        )
        u = User.objects.filter(username="newt").first()
        self.assertIsNotNone(u)
        profile = TeacherProfile.objects.get(user=u)
        self.assertEqual(profile.school, self.school)
        self.assertIn(self.subjects[0], profile.subjects.all())
        self.assertIn(self.cg, profile.classes_taught.all())

    def test_duplicate_username_rejected(self):
        self.login()
        resp = self.client.post(
            "/admin-portal/teachers/new/",
            {"first_name": "X", "last_name": "Y",
             "username": "kgosi", "password": "anotherpw1"},
        )
        self.assertContains(resp, "already taken")

    def test_short_password_rejected(self):
        self.login()
        self.client.post(
            "/admin-portal/teachers/new/",
            {"first_name": "X", "last_name": "Y",
             "username": "shorty", "password": "abc"},
        )
        self.assertFalse(User.objects.filter(username="shorty").exists())

    def test_edit_deactivates_teacher(self):
        self.login()
        # No is_active checkbox -> not active
        self.client.post(
            f"/admin-portal/teachers/{self.teacher.id}/edit/",
            {"first_name": self.teacher.user.first_name,
             "last_name": self.teacher.user.last_name,
             "email": "", "phone": "", "employee_id": "",
             "password": "", "subjects": [], "classes": []},
        )
        self.teacher.refresh_from_db()
        self.assertFalse(self.teacher.is_active)


class StudentManagementTests(AdminPortalBase):
    def test_search_by_name(self):
        self.login()
        resp = self.client.get("/admin-portal/students/?q=Stu1")
        self.assertContains(resp, "Stu1")
        self.assertNotContains(resp, "Stu2 Test")

    def test_class_filter(self):
        self.login()
        resp = self.client.get(f"/admin-portal/students/?class={self.cg.id}")
        self.assertContains(resp, "Stu1 Test")


class ClassManagementTests(AdminPortalBase):
    def test_create_class_assigns_teacher(self):
        self.login()
        resp = self.client.post(
            "/admin-portal/classes/new/",
            {"name": "Form 2A", "grade_level": "9",
             "academic_year": "2026", "class_teacher": str(self.teacher.id)},
            follow=True,
        )
        new_cg = ClassGroup.objects.get(name="Form 2A")
        self.assertEqual(new_cg.class_teacher_id, self.teacher.id)
        self.assertIn(new_cg, self.teacher.classes_taught.all())

    def test_duplicate_class_rejected(self):
        self.login()
        resp = self.client.post(
            "/admin-portal/classes/new/",
            {"name": "Form 1A", "grade_level": "8", "academic_year": "2026"},
        )
        self.assertContains(resp, "already exists")

    def test_edit_renames_class(self):
        self.login()
        self.client.post(
            f"/admin-portal/classes/{self.cg.id}/edit/",
            {"name": "Form 1B", "grade_level": "8", "academic_year": "2026"},
        )
        self.cg.refresh_from_db()
        self.assertEqual(self.cg.name, "Form 1B")


class SubjectManagementTests(AdminPortalBase):
    def test_create_subject(self):
        self.login()
        before = Subject.objects.filter(school=self.school).count()
        self.client.post("/admin-portal/subjects/",
                         {"name": "Science", "code": "SCI"}, follow=True)
        self.assertEqual(
            Subject.objects.filter(school=self.school).count(), before + 1,
        )

    def test_duplicate_code_rejected(self):
        self.login()
        resp = self.client.post("/admin-portal/subjects/",
                                {"name": "Math 2", "code": "MATH"})
        self.assertContains(resp, "already exists")


class SchoolProfileTests(AdminPortalBase):
    def test_update_basic_info(self):
        self.login()
        self.client.post(
            "/admin-portal/school/",
            {"name": self.school.name, "region": "Central",
             "address": "x", "phone": "", "email": "", "principal_name": "PP"},
        )
        self.school.refresh_from_db()
        self.assertEqual(self.school.region, "Central")
        self.assertEqual(self.school.principal_name, "PP")

    def test_empty_name_rejected(self):
        self.login()
        resp = self.client.post("/admin-portal/school/", {"name": ""})
        self.assertContains(resp, "School name is required")


class CalendarTests(AdminPortalBase):
    def test_admin_can_save_calendar(self):
        self.login()
        self.client.post(
            "/admin-portal/calendar/",
            {"year": "2098",
             "start_1": "2098-01-10", "end_1": "2098-04-05",
             "start_2": "", "end_2": "",
             "start_3": "", "end_3": ""},
        )
        self.assertTrue(
            TermSchedule.objects.filter(school=self.school, term=1, academic_year=2098).exists(),
        )


class ParentPinTests(AdminPortalBase):
    def test_parents_page_lists_phones(self):
        self.login()
        resp = self.client.get("/admin-portal/parents/")
        self.assertContains(resp, "71 222 001")

    def test_set_pin_creates_parent_account(self):
        self.login()
        self.client.post(
            "/admin-portal/parents/set-pin/",
            {"phone": "+267 71 222 001", "pin": "5678"},
        )
        u = User.objects.filter(username="26771222001", role="PARENT").first()
        self.assertIsNotNone(u)
        self.assertTrue(u.check_password("5678"))

    def test_short_pin_rejected(self):
        self.login()
        resp = self.client.post(
            "/admin-portal/parents/set-pin/",
            {"phone": "+267 71 222 001", "pin": "12"}, follow=True,
        )
        self.assertContains(resp, "at least 4 digits")
        self.assertFalse(
            User.objects.filter(username="26771222001", role="PARENT").exists(),
        )

    def test_phone_not_at_school_rejected(self):
        self.login()
        resp = self.client.post(
            "/admin-portal/parents/set-pin/",
            {"phone": "+267 99 999 999", "pin": "1234"}, follow=True,
        )
        self.assertContains(resp, "No student at this school")


class ExportsTests(AdminPortalBase):
    def setUp(self):
        for s in self.students:
            Mark.objects.create(
                student=s, subject=self.subjects[0], teacher=self.teacher,
                assessment_type="TEST", title="t1",
                score=70, max_score=100, term=1, academic_year=2026,
            )

    def test_exports_landing_renders(self):
        self.login()
        resp = self.client.get("/admin-portal/exports/")
        self.assertContains(resp, "Combined Ministry pack")
        self.assertContains(resp, "Individual exports")

    def test_marks_download(self):
        self.login()
        resp = self.client.get(
            "/admin-portal/exports/download/?kind=marks&format=xlsx&year=2026",
        )
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 3)

    def test_ministry_pack_has_all_sheets(self):
        self.login()
        resp = self.client.get("/admin-portal/exports/download/?kind=pack&year=2026")
        wb = load_workbook(BytesIO(resp.content))
        for sheet in ("Summary", "Roster", "Marks", "Attendance", "Performance"):
            self.assertIn(sheet, wb.sheetnames)

    def test_teacher_denied_exports(self):
        self.client.login(username="kgosi", password="teach1234")
        resp = self.client.get("/admin-portal/exports/")
        self.assertEqual(resp.status_code, 403)
