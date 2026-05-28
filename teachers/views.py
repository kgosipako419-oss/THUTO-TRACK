from collections import defaultdict
from datetime import date
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.timezone import now

from core.models import Attendance, BehaviorNote, ClassGroup, Enquiry, Mark, Student, Subject, TeacherProfile, Term, TermSchedule
from core.reports import build_class_term_report, build_student_term_report


def _teacher_or_403(request):
    """Return the request user's TeacherProfile or None if they don't have one."""
    return TeacherProfile.objects.select_related("school").filter(user=request.user).first()


@login_required
def dashboard(request):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)

    from django.db.models import Avg, Count, F, FloatField, Sum
    from django.db.models.functions import Cast

    year = now().year

    classes = teacher.classes_taught.select_related("school").all()
    subjects = teacher.subjects.all()

    students_qs = Student.objects.filter(
        class_group__in=classes, is_active=True,
    )
    students_count = students_qs.count()

    today_attendance_count = Attendance.objects.filter(
        recorded_by=teacher, date=date.today(),
    ).count()

    # Year attendance rate across the teacher's students
    att_year_qs = Attendance.objects.filter(student__in=students_qs, date__year=year)
    att_total = att_year_qs.count()
    att_present = att_year_qs.filter(status="P").count()
    attendance_rate = (att_present / att_total * 100) if att_total else None

    # Overall academic performance: average % across all marks this teacher recorded this year
    teacher_marks_qs = Mark.objects.filter(teacher=teacher, academic_year=year)
    teacher_marks_qs = teacher_marks_qs.annotate(
        pct=Cast(F("score"), FloatField()) * 100.0 / Cast(F("max_score"), FloatField()),
    )
    overall_stats = teacher_marks_qs.aggregate(
        avg_pct=Avg("pct"),
        assessment_count=Count("id"),
        student_count=Count("student", distinct=True),
    )

    # Per-subject performance (only subjects the teacher recorded marks for this year)
    per_subject = list(
        teacher_marks_qs.values("subject__id", "subject__name")
        .annotate(
            avg_pct=Avg("pct"),
            assessment_count=Count("id"),
            student_count=Count("student", distinct=True),
        )
        .order_by("subject__name")
    )

    # Performance band tag for the teacher header card
    avg = overall_stats.get("avg_pct")
    if avg is None:
        teacher_band = None
    elif avg >= 75:
        teacher_band = ("Excellent", "bg-emerald-100 text-emerald-800")
    elif avg >= 60:
        teacher_band = ("Good", "bg-emerald-100 text-emerald-800")
    elif avg >= 50:
        teacher_band = ("Average", "bg-amber-100 text-amber-800")
    else:
        teacher_band = ("Needs attention", "bg-red-100 text-red-800")

    recent_marks = (
        Mark.objects.filter(teacher=teacher)
        .select_related("student", "subject")
        .order_by("-recorded_at")[:10]
    )

    open_enquiries = Enquiry.objects.filter(
        from_teacher=teacher, status=Enquiry.Status.OPEN,
    ).count()

    return render(
        request,
        "teachers/dashboard.html",
        {
            "teacher": teacher,
            "year": year,
            "classes": classes,
            "subjects": subjects,
            "students_count": students_count,
            "today_attendance_count": today_attendance_count,
            "attendance_rate": attendance_rate,
            "overall_avg": overall_stats.get("avg_pct"),
            "overall_assessments": overall_stats.get("assessment_count") or 0,
            "overall_students_assessed": overall_stats.get("student_count") or 0,
            "per_subject": per_subject,
            "teacher_band": teacher_band,
            "recent_marks": recent_marks,
            "open_enquiries": open_enquiries,
        },
    )


@login_required
def class_list(request):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    classes = teacher.classes_taught.select_related("school").all()
    return render(request, "teachers/class_list.html", {"classes": classes})


@login_required
def class_detail(request, class_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    class_group = get_object_or_404(teacher.classes_taught, pk=class_id)
    students = class_group.students.filter(is_active=True)
    return render(
        request,
        "teachers/class_detail.html",
        {"class_group": class_group, "students": students},
    )


@login_required
def enter_marks(request, class_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    class_group = get_object_or_404(teacher.classes_taught, pk=class_id)
    students = class_group.students.filter(is_active=True)
    subjects = teacher.subjects.all()

    if request.method == "POST":
        subject_id = request.POST.get("subject")
        title = request.POST.get("title", "").strip()
        assessment_type = request.POST.get("assessment_type", Mark.AssessmentType.TEST)
        max_score = request.POST.get("max_score") or 100
        term = int(request.POST.get("term") or Term.TERM_1)
        academic_year = int(request.POST.get("academic_year") or now().year)

        if not subject_id or not title:
            messages.error(request, "Subject and assessment title are required.")
        else:
            created = 0
            for student in students:
                raw = request.POST.get(f"score_{student.id}")
                if raw is None or raw.strip() == "":
                    continue
                try:
                    score = float(raw)
                except ValueError:
                    continue
                Mark.objects.create(
                    student=student,
                    subject_id=subject_id,
                    teacher=teacher,
                    assessment_type=assessment_type,
                    title=title,
                    score=score,
                    max_score=max_score,
                    term=term,
                    academic_year=academic_year,
                )
                created += 1
            messages.success(request, f"Recorded {created} mark(s) for '{title}'.")
            return redirect("teachers:class_detail", class_id=class_group.id)

    return render(
        request,
        "teachers/enter_marks.html",
        {
            "class_group": class_group,
            "students": students,
            "subjects": subjects,
            "terms": Term.choices,
            "assessment_types": Mark.AssessmentType.choices,
            "current_year": now().year,
        },
    )


@login_required
def enter_attendance(request, class_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    class_group = get_object_or_404(teacher.classes_taught, pk=class_id)
    students = class_group.students.filter(is_active=True)

    if request.method == "POST":
        att_date_raw = request.POST.get("date") or date.today().isoformat()
        try:
            att_date = date.fromisoformat(att_date_raw)
        except ValueError:
            att_date = date.today()

        recorded = 0
        for student in students:
            status = request.POST.get(f"status_{student.id}")
            if not status:
                continue
            Attendance.objects.update_or_create(
                student=student,
                date=att_date,
                defaults={"status": status, "recorded_by": teacher},
            )
            recorded += 1
        messages.success(request, f"Attendance saved for {recorded} student(s) on {att_date}.")
        return redirect("teachers:class_detail", class_id=class_group.id)

    return render(
        request,
        "teachers/enter_attendance.html",
        {
            "class_group": class_group,
            "students": students,
            "today": date.today().isoformat(),
            "statuses": Attendance.Status.choices,
        },
    )


# ---------------------------------------------------------------------------
# Bulk Excel upload of students into a class
# ---------------------------------------------------------------------------

BULK_UPLOAD_COLUMNS = [
    "student_number",
    "first_name",
    "last_name",
    "gender",
    "date_of_birth",
    "parent_name",
    "parent_phone",
]


@login_required
def bulk_upload_template(request, class_id: int):
    """Return an .xlsx template the teacher can fill in and re-upload."""
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    class_group = get_object_or_404(teacher.classes_taught, pk=class_id)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="047857")
    for col, name in enumerate(BULK_UPLOAD_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill

    ws.append([
        "S-2026-100",
        "Example",
        "Student",
        "F",
        "2010-03-15",
        "Mma Example",
        "+267 71 000 000",
    ])

    instructions = (
        "Required: student_number, first_name, last_name. "
        "gender = M/F/O. date_of_birth = YYYY-MM-DD. "
        "Delete the example row before uploading."
    )
    ws.cell(row=4, column=1, value=instructions).font = Font(italic=True, color="6B7280")

    for col_letter, width in zip("ABCDEFG", (16, 16, 16, 8, 14, 24, 20)):
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"students_template_{class_group.name.replace(' ', '_')}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _parse_bulk_upload(file, class_group, teacher):
    """Read an uploaded .xlsx and return (created_students, errors).

    Errors is a list of (row_number, message) tuples. The whole upload is
    wrapped in a transaction by the caller, so a non-empty errors list means
    nothing was written.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        return [], [(0, f"Could not read file: {exc}")]

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
    except StopIteration:
        return [], [(0, "Spreadsheet is empty.")]

    missing = [c for c in ("student_number", "first_name", "last_name") if c not in header]
    if missing:
        return [], [(0, f"Missing required column(s): {', '.join(missing)}")]

    idx = {name: header.index(name) for name in BULK_UPLOAD_COLUMNS if name in header}

    created = []
    errors = []
    seen_numbers = set()

    for row_num, row in enumerate(rows, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell(name):
            i = idx.get(name)
            if i is None or i >= len(row):
                return None
            v = row[i]
            return str(v).strip() if v is not None else None

        student_number = cell("student_number")
        first_name = cell("first_name")
        last_name = cell("last_name")

        if not student_number or not first_name or not last_name:
            errors.append((row_num, "student_number, first_name, last_name are required."))
            continue

        if student_number in seen_numbers:
            errors.append((row_num, f"Duplicate student_number '{student_number}' in upload."))
            continue
        seen_numbers.add(student_number)

        if Student.objects.filter(school=class_group.school, student_number=student_number).exists():
            errors.append((row_num, f"Student '{student_number}' already exists in this school."))
            continue

        gender = (cell("gender") or "").upper()[:1]
        if gender and gender not in {"M", "F", "O"}:
            errors.append((row_num, f"Invalid gender '{gender}'. Use M, F or O."))
            continue

        dob = None
        dob_raw = cell("date_of_birth")
        if dob_raw:
            try:
                dob = date.fromisoformat(dob_raw[:10])
            except ValueError:
                errors.append((row_num, f"Invalid date_of_birth '{dob_raw}'. Use YYYY-MM-DD."))
                continue

        created.append(
            Student(
                school=class_group.school,
                student_number=student_number,
                first_name=first_name,
                last_name=last_name,
                gender=gender,
                date_of_birth=dob,
                class_group=class_group,
                parent_name=cell("parent_name") or "",
                parent_phone=cell("parent_phone") or "",
            )
        )

    return created, errors


@login_required
def bulk_upload_students(request, class_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    class_group = get_object_or_404(teacher.classes_taught.select_related("school"), pk=class_id)

    errors = []
    created_count = 0

    if request.method == "POST":
        uploaded = request.FILES.get("file")
        if not uploaded:
            errors = [(0, "Please choose an .xlsx file to upload.")]
        elif not uploaded.name.lower().endswith(".xlsx"):
            errors = [(0, "File must be an .xlsx workbook.")]
        else:
            to_create, errors = _parse_bulk_upload(uploaded, class_group, teacher)
            if not errors:
                try:
                    with transaction.atomic():
                        Student.objects.bulk_create(to_create)
                    created_count = len(to_create)
                    messages.success(
                        request,
                        f"Added {created_count} student(s) to {class_group.name}.",
                    )
                    return redirect("teachers:class_detail", class_id=class_group.id)
                except IntegrityError as exc:
                    errors = [(0, f"Database rejected upload: {exc}")]

    return render(
        request,
        "teachers/bulk_upload.html",
        {
            "class_group": class_group,
            "errors": errors,
            "columns": BULK_UPLOAD_COLUMNS,
        },
    )


# ---------------------------------------------------------------------------
# Student profile (full history view)
# ---------------------------------------------------------------------------

@login_required
def student_detail(request, student_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)

    student = get_object_or_404(
        Student.objects.select_related("class_group", "school"),
        pk=student_id,
        school=teacher.school,
        class_group__in=teacher.classes_taught.all(),
    )

    year = now().year

    marks_qs = (
        Mark.objects.filter(student=student)
        .select_related("subject", "teacher__user")
        .order_by("subject__name", "term", "-recorded_at")
    )
    marks_by_subject = defaultdict(list)
    for m in marks_qs:
        marks_by_subject[m.subject].append(m)

    subject_averages = []
    for subject, marks in marks_by_subject.items():
        avg = sum(m.percentage for m in marks) / len(marks)
        subject_averages.append({"subject": subject, "average": avg, "count": len(marks)})
    subject_averages.sort(key=lambda r: r["subject"].name)

    attendance_qs = Attendance.objects.filter(student=student, date__year=year)
    attendance_counts = {label: 0 for _, label in Attendance.Status.choices}
    status_to_label = dict(Attendance.Status.choices)
    for a in attendance_qs:
        attendance_counts[status_to_label[a.status]] += 1
    total_recorded = sum(attendance_counts.values())
    present_count = attendance_counts.get("Present", 0)
    attendance_rate = (present_count / total_recorded * 100) if total_recorded else None

    recent_attendance = Attendance.objects.filter(student=student).order_by("-date")[:10]
    behavior_notes = (
        BehaviorNote.objects.filter(student=student)
        .select_related("teacher__user")
        .order_by("-recorded_at")[:20]
    )

    return render(
        request,
        "teachers/student_detail.html",
        {
            "student": student,
            "marks_by_subject": dict(marks_by_subject),
            "subject_averages": subject_averages,
            "attendance_counts": attendance_counts,
            "attendance_rate": attendance_rate,
            "total_recorded": total_recorded,
            "recent_attendance": recent_attendance,
            "behavior_notes": behavior_notes,
            "year": year,
            "terms": Term.choices,
        },
    )


# ---------------------------------------------------------------------------
# PDF term reports
# ---------------------------------------------------------------------------

def _parse_term_year(request, default_year):
    try:
        term = int(request.GET.get("term") or 1)
    except ValueError:
        term = 1
    if term not in (1, 2, 3):
        term = 1
    try:
        year = int(request.GET.get("year") or default_year)
    except ValueError:
        year = default_year
    return term, year


def _pdf_response(content: bytes, filename: str) -> HttpResponse:
    response = HttpResponse(content, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


@login_required
def student_term_report(request, student_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    student = get_object_or_404(
        Student.objects.select_related("class_group", "school"),
        pk=student_id,
        school=teacher.school,
        class_group__in=teacher.classes_taught.all(),
    )
    term, year = _parse_term_year(request, now().year)
    pdf = build_student_term_report(student, term, year)
    safe = student.full_name.replace(" ", "_")
    return _pdf_response(pdf, f"{safe}_Term{term}_{year}.pdf")


@login_required
def class_term_reports(request, class_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    class_group = get_object_or_404(
        teacher.classes_taught.select_related("school"),
        pk=class_id,
    )
    term, year = _parse_term_year(request, now().year)
    pdf = build_class_term_report(class_group, term, year)
    safe = class_group.name.replace(" ", "_")
    return _pdf_response(pdf, f"{safe}_Term{term}_{year}_reports.pdf")


# ---------------------------------------------------------------------------
# Behavior notes (add)
# ---------------------------------------------------------------------------

@login_required
def add_behavior_note(request, student_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    student = get_object_or_404(
        Student.objects.select_related("class_group", "school"),
        pk=student_id,
        school=teacher.school,
        class_group__in=teacher.classes_taught.all(),
    )

    if request.method == "POST":
        category = request.POST.get("category") or BehaviorNote.Category.NEUTRAL
        note_text = (request.POST.get("note") or "").strip()
        if category not in dict(BehaviorNote.Category.choices):
            messages.error(request, "Invalid category.")
        elif not note_text:
            messages.error(request, "Note text is required.")
        else:
            BehaviorNote.objects.create(
                student=student, teacher=teacher, category=category, note=note_text,
            )
            messages.success(request, "Behavior note saved.")
            return redirect("teachers:student_detail", student_id=student.id)

    return render(
        request,
        "teachers/add_behavior_note.html",
        {"student": student, "categories": BehaviorNote.Category.choices},
    )


# ---------------------------------------------------------------------------
# Subject management (list + create)
# ---------------------------------------------------------------------------

@login_required
def subjects_manage(request):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        code = (request.POST.get("code") or "").strip().upper()
        if not name or not code:
            messages.error(request, "Name and code are required.")
        elif Subject.objects.filter(school=teacher.school, code=code).exists():
            messages.error(request, f"A subject with code '{code}' already exists for this school.")
        else:
            subject = Subject.objects.create(school=teacher.school, name=name, code=code)
            teacher.subjects.add(subject)
            messages.success(request, f"Added subject '{name}' and assigned it to you.")
            return redirect("teachers:subjects_manage")

    subjects = Subject.objects.filter(school=teacher.school).order_by("name")
    my_subject_ids = set(teacher.subjects.values_list("id", flat=True))
    return render(
        request,
        "teachers/subjects_manage.html",
        {"subjects": subjects, "my_subject_ids": my_subject_ids, "teacher": teacher},
    )


# ---------------------------------------------------------------------------
# Class management (create new — list reuses existing class_list view)
# ---------------------------------------------------------------------------

@login_required
def class_create(request):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)

    current_year = now().year

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        try:
            grade_level = int(request.POST.get("grade_level") or 0)
        except ValueError:
            grade_level = 0
        try:
            academic_year = int(request.POST.get("academic_year") or current_year)
        except ValueError:
            academic_year = current_year

        if not name or grade_level <= 0:
            messages.error(request, "Name and a valid grade level are required.")
        elif ClassGroup.objects.filter(
            school=teacher.school, name=name, academic_year=academic_year
        ).exists():
            messages.error(
                request,
                f"A class named '{name}' already exists for {academic_year}.",
            )
        else:
            class_group = ClassGroup.objects.create(
                school=teacher.school,
                name=name,
                grade_level=grade_level,
                academic_year=academic_year,
                class_teacher=teacher,
            )
            teacher.classes_taught.add(class_group)
            messages.success(request, f"Created class '{name}' and assigned you as class teacher.")
            return redirect("teachers:class_detail", class_id=class_group.id)

    return render(
        request,
        "teachers/class_create.html",
        {"current_year": current_year},
    )


# ---------------------------------------------------------------------------
# School calendar: term date ranges
# ---------------------------------------------------------------------------

def _parse_date(raw):
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw[:10])
    except ValueError:
        return "INVALID"


@login_required
def school_calendar(request):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)

    current_year = now().year
    try:
        year = int(request.GET.get("year") or current_year)
    except ValueError:
        year = current_year

    errors_per_term = {}

    if request.method == "POST":
        try:
            year = int(request.POST.get("year") or current_year)
        except ValueError:
            year = current_year

        row_changes = []
        for term_num in (1, 2, 3):
            start = _parse_date(request.POST.get(f"start_{term_num}"))
            end = _parse_date(request.POST.get(f"end_{term_num}"))

            if start == "INVALID" or end == "INVALID":
                errors_per_term[term_num] = "Invalid date format."
                continue

            if start is None and end is None:
                row_changes.append(("clear", term_num))
                continue

            if start is None or end is None:
                errors_per_term[term_num] = "Both start and end dates are required."
                continue

            if end < start:
                errors_per_term[term_num] = "End date must be on or after start date."
                continue

            row_changes.append(("upsert", term_num, start, end))

        if not errors_per_term:
            with transaction.atomic():
                for change in row_changes:
                    if change[0] == "clear":
                        TermSchedule.objects.filter(
                            school=teacher.school, academic_year=year, term=change[1],
                        ).delete()
                    else:
                        _, term_num, start, end = change
                        TermSchedule.objects.update_or_create(
                            school=teacher.school,
                            academic_year=year,
                            term=term_num,
                            defaults={"start_date": start, "end_date": end},
                        )
            messages.success(request, f"Calendar for {year} saved.")
            return redirect(f"{request.path}?year={year}")

    existing = {
        ts.term: ts
        for ts in TermSchedule.objects.filter(school=teacher.school, academic_year=year)
    }

    term_rows = [
        {
            "num": n,
            "schedule": existing.get(n),
            "error": errors_per_term.get(n),
        }
        for n in (1, 2, 3)
    ]

    return render(
        request,
        "teachers/school_calendar.html",
        {
            "year": year,
            "prev_year": year - 1,
            "next_year": year + 1,
            "current_year": current_year,
            "term_rows": term_rows,
            "teacher": teacher,
        },
    )


# ---------------------------------------------------------------------------
# Single-student CRUD: create, edit, delete
# ---------------------------------------------------------------------------

def _student_from_form(request, student=None):
    """Populate a Student instance from POST data. Returns (instance, errors_dict)."""
    errors = {}
    first_name = (request.POST.get("first_name") or "").strip()
    last_name = (request.POST.get("last_name") or "").strip()
    student_number = (request.POST.get("student_number") or "").strip()
    if not first_name:
        errors["first_name"] = "First name is required."
    if not last_name:
        errors["last_name"] = "Last name is required."
    if not student_number:
        errors["student_number"] = "Student number is required."

    dob_raw = (request.POST.get("date_of_birth") or "").strip()
    dob = None
    if dob_raw:
        try:
            dob = date.fromisoformat(dob_raw[:10])
        except ValueError:
            errors["date_of_birth"] = "Use the format YYYY-MM-DD."

    gender = (request.POST.get("gender") or "").upper()[:1]
    if gender and gender not in {"M", "F", "O"}:
        errors["gender"] = "Choose Male, Female, or Other."

    if student is None:
        student = Student()
    student.first_name = first_name
    student.last_name = last_name
    student.student_number = student_number
    student.date_of_birth = dob
    student.gender = gender
    student.parent_name = (request.POST.get("parent_name") or "").strip()
    student.parent_phone = (request.POST.get("parent_phone") or "").strip()
    return student, errors


@login_required
def student_create(request, class_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    class_group = get_object_or_404(
        teacher.classes_taught.select_related("school"),
        pk=class_id,
    )

    if request.method == "POST":
        student, errors = _student_from_form(request)
        if not errors:
            if Student.objects.filter(
                school=class_group.school, student_number=student.student_number,
            ).exists():
                errors["student_number"] = (
                    f"Student number '{student.student_number}' already exists in this school."
                )
        if not errors:
            student.school = class_group.school
            student.class_group = class_group
            student.save()
            messages.success(request, f"Added {student.full_name} to {class_group.name}.")
            return redirect("teachers:student_detail", student_id=student.id)
    else:
        student = Student(class_group=class_group, school=class_group.school)
        errors = {}

    return render(
        request,
        "teachers/student_form.html",
        {
            "student": student,
            "errors": errors,
            "class_group": class_group,
            "mode": "create",
            "genders": Student.Gender.choices,
        },
    )


@login_required
def student_edit(request, student_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    student = get_object_or_404(
        Student.objects.select_related("class_group", "school"),
        pk=student_id,
        school=teacher.school,
        class_group__in=teacher.classes_taught.all(),
    )

    teacher_classes = teacher.classes_taught.select_related("school").all()

    if request.method == "POST":
        student, errors = _student_from_form(request, student=student)
        new_class_id = request.POST.get("class_group")
        try:
            new_class_id = int(new_class_id) if new_class_id else student.class_group_id
        except ValueError:
            new_class_id = student.class_group_id
        if not teacher_classes.filter(pk=new_class_id).exists():
            errors["class_group"] = "You can only move the student into a class you teach."
        if not errors:
            duplicate = (
                Student.objects.filter(
                    school=student.school, student_number=student.student_number,
                )
                .exclude(pk=student.pk)
                .exists()
            )
            if duplicate:
                errors["student_number"] = (
                    f"Student number '{student.student_number}' already exists in this school."
                )
        if not errors:
            student.class_group_id = new_class_id
            student.is_active = bool(request.POST.get("is_active"))
            student.save()
            messages.success(request, f"Updated {student.full_name}.")
            return redirect("teachers:student_detail", student_id=student.id)
    else:
        errors = {}

    return render(
        request,
        "teachers/student_form.html",
        {
            "student": student,
            "errors": errors,
            "class_group": student.class_group,
            "teacher_classes": teacher_classes,
            "mode": "edit",
            "genders": Student.Gender.choices,
        },
    )


@login_required
def student_delete(request, student_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    student = get_object_or_404(
        Student.objects.select_related("class_group", "school"),
        pk=student_id,
        school=teacher.school,
        class_group__in=teacher.classes_taught.all(),
    )

    impact = {
        "marks": student.marks.count(),
        "attendance": student.attendance_records.count(),
        "behavior_notes": student.behavior_notes.count(),
    }

    if request.method == "POST":
        confirm = (request.POST.get("confirm") or "").strip()
        if confirm != student.student_number:
            messages.error(
                request,
                "Confirmation did not match the student number. Nothing was deleted.",
            )
        else:
            class_id = student.class_group_id
            name = student.full_name
            student.delete()
            messages.success(request, f"Deleted {name} and all their records.")
            return redirect("teachers:class_detail", class_id=class_id)

    return render(
        request,
        "teachers/student_delete.html",
        {"student": student, "impact": impact},
    )


# ---------------------------------------------------------------------------
# Enquiries to admin / HR
# ---------------------------------------------------------------------------

@login_required
def enquiry_list(request):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    enquiries = Enquiry.objects.filter(from_teacher=teacher)
    open_count = enquiries.filter(status=Enquiry.Status.OPEN).count()
    return render(
        request,
        "teachers/enquiry_list.html",
        {"enquiries": enquiries, "open_count": open_count},
    )


@login_required
def enquiry_create(request):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)

    errors = {}
    if request.method == "POST":
        subject = (request.POST.get("subject") or "").strip()
        body = (request.POST.get("body") or "").strip()
        category = request.POST.get("category") or Enquiry.Category.OTHER
        if not subject:
            errors["subject"] = "Subject is required."
        if not body:
            errors["body"] = "Please describe your enquiry."
        if category not in dict(Enquiry.Category.choices):
            errors["category"] = "Choose a valid category."
        if not errors:
            enquiry = Enquiry.objects.create(
                school=teacher.school,
                from_teacher=teacher,
                category=category,
                subject=subject,
                body=body,
            )
            messages.success(request, "Enquiry sent to admin / HR.")
            return redirect("teachers:enquiry_detail", enquiry_id=enquiry.id)

    return render(
        request,
        "teachers/enquiry_form.html",
        {
            "categories": Enquiry.Category.choices,
            "values": request.POST if request.method == "POST" else {},
            "errors": errors,
        },
    )


@login_required
def enquiry_detail(request, enquiry_id: int):
    teacher = _teacher_or_403(request)
    if not teacher:
        return render(request, "teachers/no_profile.html", status=403)
    enquiry = get_object_or_404(Enquiry, pk=enquiry_id, from_teacher=teacher)
    return render(request, "teachers/enquiry_detail.html", {"enquiry": enquiry})
