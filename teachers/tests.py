"""End-to-end tests for the teacher portal."""

from datetime import date
from io import BytesIO

from django.test import TestCase
from openpyxl import Workbook, load_workbook

from core.factories import (
    make_class,
    make_school,
    make_student,
    make_subject,
    make_teacher,
)
from core.models import Attendance, BehaviorNote, Mark, Student


class TeacherPortalBase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.school = make_school()
        cls.teacher = make_teacher(cls.school, username="kgosi", password="teach1234")
        cls.subjects = [
            make_subject(cls.school, code="MATH", name="Mathematics"),
            make_subject(cls.school, code="ENG", name="English"),
        ]
        cls.teacher.subjects.set(cls.subjects)
        cls.cg = make_class(cls.school, class_teacher=cls.teacher)
        cls.teacher.classes_taught.add(cls.cg)
        cls.students = [
            make_student(cls.school, cls.cg, student_number=f"S-{i:03d}",
                         first_name=f"Stu{i}", last_name="Test",
                         parent_phone=f"+267 71 000 {i:03d}")
            for i in range(1, 4)
        ]

    def login(self):
        self.assertTrue(self.client.login(username="kgosi", password="teach1234"))


class AuthAndDashboardTests(TeacherPortalBase):
    def test_login_required_for_dashboard(self):
        resp = self.client.get("/teachers/")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login/", resp["Location"])

    def test_dashboard_renders_for_teacher(self):
        self.login()
        resp = self.client.get("/teachers/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Form 1A")
        self.assertContains(resp, "My classes")

    def test_user_without_teacher_profile_gets_no_profile_screen(self):
        from core.factories import make_user
        make_user("stranger", "strangerpw1", role="TEACHER")
        self.assertTrue(self.client.login(username="stranger", password="strangerpw1"))
        resp = self.client.get("/teachers/")
        self.assertEqual(resp.status_code, 403)


class ClassDetailTests(TeacherPortalBase):
    def test_class_detail_lists_students(self):
        self.login()
        resp = self.client.get(f"/teachers/classes/{self.cg.id}/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Stu1 Test")
        self.assertContains(resp, "S-001")

    def test_foreign_class_404(self):
        other_school = make_school(name="Other", code="OTH-1")
        other_cg = make_class(other_school)
        self.login()
        resp = self.client.get(f"/teachers/classes/{other_cg.id}/")
        self.assertEqual(resp.status_code, 404)


class MarksEntryTests(TeacherPortalBase):
    def test_enter_marks_persists_rows(self):
        self.login()
        data = {
            "subject": str(self.subjects[0].id),
            "title": "Mid-term test",
            "assessment_type": "TEST",
            "max_score": "100",
            "term": "1",
            "academic_year": "2026",
        }
        for s in self.students:
            data[f"score_{s.id}"] = "80"
        resp = self.client.post(f"/teachers/classes/{self.cg.id}/marks/", data, follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            Mark.objects.filter(teacher=self.teacher, title="Mid-term test").count(),
            3,
        )

    def test_missing_subject_rejected(self):
        self.login()
        resp = self.client.post(
            f"/teachers/classes/{self.cg.id}/marks/",
            {"title": "x", "subject": ""},
        )
        # form re-rendered, no marks created
        self.assertEqual(Mark.objects.count(), 0)


class AttendanceEntryTests(TeacherPortalBase):
    def test_attendance_recorded(self):
        self.login()
        data = {"date": "2026-05-29"}
        for s in self.students:
            data[f"status_{s.id}"] = "P"
        resp = self.client.post(
            f"/teachers/classes/{self.cg.id}/attendance/", data, follow=True,
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            Attendance.objects.filter(student__class_group=self.cg).count(), 3,
        )

    def test_attendance_is_idempotent(self):
        self.login()
        data = {"date": "2026-05-29"}
        for s in self.students:
            data[f"status_{s.id}"] = "P"
        self.client.post(f"/teachers/classes/{self.cg.id}/attendance/", data)
        # Second post with different status -> updates, no duplicates
        data2 = {**data, f"status_{self.students[0].id}": "A"}
        self.client.post(f"/teachers/classes/{self.cg.id}/attendance/", data2)
        self.assertEqual(Attendance.objects.filter(student=self.students[0]).count(), 1)
        self.assertEqual(
            Attendance.objects.get(student=self.students[0]).status, "A",
        )


class BehaviorNoteTests(TeacherPortalBase):
    def test_add_behavior_note_creates_row(self):
        self.login()
        s = self.students[0]
        resp = self.client.post(
            f"/teachers/students/{s.id}/behavior/add/",
            {"category": "POS", "note": "Great work!"}, follow=True,
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(BehaviorNote.objects.filter(student=s, note="Great work!").exists())

    def test_empty_note_rejected(self):
        self.login()
        s = self.students[0]
        resp = self.client.post(
            f"/teachers/students/{s.id}/behavior/add/",
            {"category": "POS", "note": "   "},
        )
        self.assertContains(resp, "Note text is required")
        self.assertFalse(BehaviorNote.objects.filter(student=s).exists())


class StudentCrudTests(TeacherPortalBase):
    def test_create_student(self):
        self.login()
        before = Student.objects.count()
        resp = self.client.post(
            f"/teachers/classes/{self.cg.id}/students/new/",
            {
                "first_name": "New", "last_name": "Kid",
                "student_number": "NEW-1",
                "date_of_birth": "2012-01-01",
                "gender": "M",
                "parent_name": "Parent",
                "parent_phone": "+267 71 999 999",
            }, follow=True,
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(Student.objects.count(), before + 1)

    def test_duplicate_student_number_rejected(self):
        self.login()
        resp = self.client.post(
            f"/teachers/classes/{self.cg.id}/students/new/",
            {"first_name": "Dup", "last_name": "X", "student_number": "S-001"},
        )
        self.assertContains(resp, "already exists")

    def test_edit_student_updates_fields(self):
        self.login()
        s = self.students[0]
        resp = self.client.post(
            f"/teachers/students/{s.id}/edit/",
            {
                "first_name": "Renamed", "last_name": s.last_name,
                "student_number": s.student_number,
                "class_group": str(self.cg.id),
                "is_active": "on",
            }, follow=True,
        )
        s.refresh_from_db()
        self.assertEqual(s.first_name, "Renamed")

    def test_delete_requires_correct_confirmation(self):
        self.login()
        s = self.students[0]
        self.client.post(f"/teachers/students/{s.id}/delete/", {"confirm": "wrong"})
        self.assertTrue(Student.objects.filter(id=s.id).exists())
        self.client.post(f"/teachers/students/{s.id}/delete/", {"confirm": s.student_number})
        self.assertFalse(Student.objects.filter(id=s.id).exists())

    def test_delete_cascades_marks_attendance_notes(self):
        self.login()
        s = self.students[0]
        Mark.objects.create(
            student=s, subject=self.subjects[0], teacher=self.teacher,
            assessment_type="TEST", title="t", score=50, max_score=100,
            term=1, academic_year=2026,
        )
        Attendance.objects.create(student=s, date=date(2026, 5, 29), status="P",
                                  recorded_by=self.teacher)
        BehaviorNote.objects.create(student=s, teacher=self.teacher,
                                    category="POS", note="n")
        sid = s.id
        self.client.post(f"/teachers/students/{s.id}/delete/", {"confirm": s.student_number})
        self.assertFalse(Mark.objects.filter(student_id=sid).exists())
        self.assertFalse(Attendance.objects.filter(student_id=sid).exists())
        self.assertFalse(BehaviorNote.objects.filter(student_id=sid).exists())


class BulkUploadTests(TeacherPortalBase):
    def _xlsx(self, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(["student_number", "first_name", "last_name", "gender",
                   "date_of_birth", "parent_name", "parent_phone"])
        for r in rows:
            ws.append(r)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "upload.xlsx"
        return buf

    def test_template_download(self):
        self.login()
        resp = self.client.get(f"/teachers/classes/{self.cg.id}/upload/template/")
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        self.assertIn("student_number", [c.value for c in next(wb.active.iter_rows(max_row=1))])

    def test_good_upload_creates_students(self):
        self.login()
        before = Student.objects.count()
        xlsx = self._xlsx([
            ["NEW-1", "A", "Apple", "M", "2012-01-01", "P", "+267 71 1 1 1"],
            ["NEW-2", "B", "Berry", "F", "", "", ""],
        ])
        self.client.post(f"/teachers/classes/{self.cg.id}/upload/", {"file": xlsx}, follow=True)
        self.assertEqual(Student.objects.count(), before + 2)

    def test_bad_row_rejects_whole_batch(self):
        self.login()
        before = Student.objects.count()
        xlsx = self._xlsx([
            ["NEW-1", "A", "Apple", "M", "", "", ""],
            ["", "Bad", "Row", "", "", "", ""],  # missing student_number
        ])
        resp = self.client.post(f"/teachers/classes/{self.cg.id}/upload/", {"file": xlsx})
        self.assertContains(resp, "Upload rejected")
        self.assertEqual(Student.objects.count(), before)


class PdfReportTests(TeacherPortalBase):
    def test_student_term_report_returns_pdf(self):
        self.login()
        Mark.objects.create(
            student=self.students[0], subject=self.subjects[0], teacher=self.teacher,
            assessment_type="TEST", title="T1", score=80, max_score=100,
            term=1, academic_year=2026,
        )
        resp = self.client.get(
            f"/teachers/students/{self.students[0].id}/report/?term=1&year=2026",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_class_term_report_combined(self):
        self.login()
        for s in self.students:
            Mark.objects.create(
                student=s, subject=self.subjects[0], teacher=self.teacher,
                assessment_type="TEST", title="T1", score=80, max_score=100,
                term=1, academic_year=2026,
            )
        resp = self.client.get(
            f"/teachers/classes/{self.cg.id}/reports/?term=1&year=2026",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.content.startswith(b"%PDF"))


class SubjectClassManagementTests(TeacherPortalBase):
    def test_teacher_creates_subject(self):
        self.login()
        before = self.teacher.subjects.count()
        self.client.post("/teachers/subjects/", {"name": "History", "code": "HIST"}, follow=True)
        self.assertEqual(self.teacher.subjects.count(), before + 1)

    def test_teacher_creates_class(self):
        self.login()
        resp = self.client.post(
            "/teachers/classes/new/",
            {"name": "Form 2B", "grade_level": "9", "academic_year": "2026"},
            follow=True,
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(self.teacher.classes_taught.filter(name="Form 2B").exists())


class EnquiryFlowTests(TeacherPortalBase):
    def test_teacher_creates_enquiry_and_sees_it(self):
        self.login()
        resp = self.client.post(
            "/teachers/enquiries/new/",
            {"category": "HR", "subject": "Leave request",
             "body": "Need a day off"}, follow=True,
        )
        self.assertContains(resp, "Leave request")
        from core.models import Enquiry
        self.assertEqual(
            Enquiry.objects.filter(from_teacher=self.teacher).count(), 1,
        )


class SchoolCalendarTests(TeacherPortalBase):
    def test_calendar_creates_terms(self):
        self.login()
        from core.models import TermSchedule
        resp = self.client.post(
            "/teachers/calendar/",
            {
                "year": "2099",
                "start_1": "2099-01-10", "end_1": "2099-04-05",
                "start_2": "2099-05-01", "end_2": "2099-08-04",
                "start_3": "2099-09-01", "end_3": "2099-12-03",
            }, follow=True,
        )
        self.assertEqual(
            TermSchedule.objects.filter(school=self.school, academic_year=2099).count(), 3,
        )

    def test_end_before_start_rejected(self):
        self.login()
        from core.models import TermSchedule
        resp = self.client.post(
            "/teachers/calendar/",
            {
                "year": "2099",
                "start_1": "2099-04-01", "end_1": "2099-01-01",
                "start_2": "", "end_2": "",
                "start_3": "", "end_3": "",
            },
        )
        self.assertContains(resp, "End date must be on or after start date")
        self.assertFalse(
            TermSchedule.objects.filter(school=self.school, academic_year=2099).exists(),
        )
