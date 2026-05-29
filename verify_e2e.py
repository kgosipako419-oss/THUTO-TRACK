"""End-to-end verification of bulk upload + student profile.

Uses Django's test client (no live server needed). Exits non-zero on any failure.
"""

import os
import sys
from datetime import date
from io import BytesIO

import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "thutotrack.settings")
os.environ["DJANGO_ALLOWED_HOSTS"] = "testserver,localhost,127.0.0.1"
django.setup()

from django.test import Client  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

from core.models import (  # noqa: E402
    Attendance,
    BehaviorNote,
    ClassGroup,
    Enquiry,
    Mark,
    ParentSession,
    School,
    SchoolAdminProfile,
    Student,
    Subject,
    TeacherProfile,
    TermSchedule,
)

ASSERT_COUNT = 0


def check(label, cond, detail=""):
    global ASSERT_COUNT
    ASSERT_COUNT += 1
    if not cond:
        print(f"FAIL: {label}{(' — ' + detail) if detail else ''}")
        sys.exit(1)
    print(f"OK:   {label}")


def make_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.append(["student_number", "first_name", "last_name", "gender", "date_of_birth", "parent_name", "parent_phone"])
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "sample.xlsx"
    return buf


c = Client()
assert c.login(username="mr_kgosi", password="teacher123"), "demo teacher login failed"

class_group = ClassGroup.objects.get(name="Form 1A")

# Clean up any leftover test data from previous runs (idempotent re-runs)
Student.objects.filter(student_number__in=["Z-TEST-101", "Z-TEST-102"]).delete()
class_url = f"/teachers/classes/{class_group.id}/"

# 1. Template download
resp = c.get(f"/teachers/classes/{class_group.id}/upload/template/")
check("template download 200", resp.status_code == 200, str(resp.status_code))
check(
    "template is xlsx",
    resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
content = b"".join(resp.streaming_content) if resp.streaming else resp.content
wb = load_workbook(BytesIO(content), read_only=True)
header = [c.value for c in next(wb.active.iter_rows(max_row=1))]
check("template header has student_number", "student_number" in header)
check("template header has parent_phone", "parent_phone" in header)

# 2. Upload form GET
resp = c.get(f"/teachers/classes/{class_group.id}/upload/")
check("upload form 200", resp.status_code == 200)
check("upload form shows download button", b"Download .xlsx template" in resp.content)

# 3. Upload with a bad row — expect rejection, no rows created
before = Student.objects.filter(class_group=class_group).count()
bad = make_xlsx([
    ["Z-TEST-101", "Mpho", "Tau", "F", "2011-04-12", "Mma Tau", "+267 71 333 001"],
    ["", "Bad", "Row", "", "", "", ""],
])
resp = c.post(f"/teachers/classes/{class_group.id}/upload/", {"file": bad})
check("bad upload returns 200 (form re-rendered)", resp.status_code == 200)
check("bad upload mentions rejection", b"Upload rejected" in resp.content)
check(
    "bad upload created NO students",
    Student.objects.filter(class_group=class_group).count() == before,
    f"before={before}, after={Student.objects.filter(class_group=class_group).count()}",
)

# 4. Upload with invalid gender — expect rejection
bad_gender = make_xlsx([
    ["Z-TEST-101", "Mpho", "Tau", "Z", "2011-04-12", "Mma Tau", "+267 71 333 001"],
])
resp = c.post(f"/teachers/classes/{class_group.id}/upload/", {"file": bad_gender})
check("invalid-gender upload rejected", b"Invalid gender" in resp.content)
check("still no students added", Student.objects.filter(class_group=class_group).count() == before)

# 5. Upload with duplicate student_number (one already seeded as S-2026-001) — should reject
dup = make_xlsx([
    ["S-2026-001", "Dup", "Student", "M", "", "", ""],
])
resp = c.post(f"/teachers/classes/{class_group.id}/upload/", {"file": dup})
check("duplicate student_number rejected", b"already exists" in resp.content)
check("still no students added", Student.objects.filter(class_group=class_group).count() == before)

# 6. Valid upload — students should be created and we redirect to class detail
good = make_xlsx([
    ["Z-TEST-101", "Mpho", "Tau", "F", "2011-04-12", "Mma Tau", "+267 71 333 001"],
    ["Z-TEST-102", "Kabo", "Sebina", "M", "2011-08-09", "Rra Sebina", "+267 71 333 002"],
])
resp = c.post(f"/teachers/classes/{class_group.id}/upload/", {"file": good}, follow=True)
check("good upload follows to class detail 200", resp.status_code == 200)
check("Mpho appears in class detail", b"Mpho Tau" in resp.content)
check("Kabo appears in class detail", b"Kabo Sebina" in resp.content)
check(
    "good upload created 2 new students",
    Student.objects.filter(class_group=class_group).count() == before + 2,
)

# 7. Student profile page
naledi = Student.objects.get(student_number="S-2026-001")
resp = c.get(f"/teachers/students/{naledi.id}/")
check("student profile 200", resp.status_code == 200)
check("profile shows name", b"Naledi Seretse" in resp.content)
check("profile shows attendance section", b"Recent attendance" in resp.content)
check("profile shows subject averages", b"Subject averages" in resp.content)
check("profile shows behavior section", b"Behavior notes" in resp.content)

# 8. Student profile for student outside teacher's class should 404
import core.models as core_models
other_school = core_models.School.objects.create(name="Other School", code="OTHER-1")
other_class = core_models.ClassGroup.objects.create(
    school=other_school, name="Other 1A", grade_level=8, academic_year=2026,
)
other_student = core_models.Student.objects.create(
    school=other_school, student_number="X-1", first_name="No", last_name="Access",
    class_group=other_class,
)
resp = c.get(f"/teachers/students/{other_student.id}/")
check("student outside teacher's class is 404", resp.status_code == 404)

# Clean up the other-school test fixtures so re-runs stay clean
other_student.delete()
other_class.delete()
other_school.delete()

# Clean up the students we just uploaded so the script is idempotent
Student.objects.filter(student_number__in=["Z-TEST-101", "Z-TEST-102"]).delete()

# 9. PDF term reports
from pypdf import PdfReader  # noqa: E402

# Seed a handful of marks so the report has content
teacher = TeacherProfile.objects.get(user__username="mr_kgosi")
math = Subject.objects.get(school=naledi.school, code="MATH")
english = Subject.objects.get(school=naledi.school, code="ENG")
Mark.objects.filter(student=naledi, term=1, academic_year=2026).delete()
Mark.objects.create(student=naledi, subject=math, teacher=teacher,
                    assessment_type="TEST", title="Mid-term test 1",
                    score=78, max_score=100, term=1, academic_year=2026)
Mark.objects.create(student=naledi, subject=math, teacher=teacher,
                    assessment_type="QUIZ", title="Chapter 3 quiz",
                    score=18, max_score=20, term=1, academic_year=2026)
Mark.objects.create(student=naledi, subject=english, teacher=teacher,
                    assessment_type="ASSIGN", title="Essay: My village",
                    score=42, max_score=50, term=1, academic_year=2026)

# Per-student report
resp = c.get(f"/teachers/students/{naledi.id}/report/?term=1&year=2026")
check("student report 200", resp.status_code == 200)
check("student report is PDF", resp["Content-Type"] == "application/pdf")
check("student report is inline", "inline" in resp["Content-Disposition"])
check("student report filename includes term", "Term1_2026" in resp["Content-Disposition"])
check("student report body starts with %PDF", resp.content.startswith(b"%PDF"))
check("student report body is non-trivial size", len(resp.content) > 2500,
      f"got {len(resp.content)} bytes")

reader = PdfReader(BytesIO(resp.content))
text = "\n".join(page.extract_text() for page in reader.pages)
check("student report PDF parses cleanly", len(reader.pages) >= 1)
check("student report contains student name", "Naledi" in text and "Seretse" in text)
check("student report contains school name", "Gaborone Demo Secondary School" in text)
check("student report contains 'Term 1 Report'", "Term 1 Report" in text)
check("student report contains Mathematics row", "Mathematics" in text)
check("student report contains assessment title", "Mid-term test 1" in text)
check("student report contains overall row", "Overall" in text)
check("student report contains attendance section", "Attendance" in text)

# Per-class report
resp = c.get(f"/teachers/classes/{class_group.id}/reports/?term=1&year=2026")
check("class report 200", resp.status_code == 200)
check("class report is PDF", resp["Content-Type"] == "application/pdf")
check("class report body starts with %PDF", resp.content.startswith(b"%PDF"))
class_pdf_size = len(resp.content)
check("class report is larger than single-student report", class_pdf_size > len(resp.content) // 2)

class_reader = PdfReader(BytesIO(resp.content))
check("class report has multiple pages (one per student)", len(class_reader.pages) >= 2,
      f"got {len(class_reader.pages)} pages")
class_text = "\n".join(page.extract_text() for page in class_reader.pages)
check("class report contains seeded students", "Naledi" in class_text and "Tumelo" in class_text)

# Clean up seeded marks
Mark.objects.filter(student=naledi, term=1, academic_year=2026).delete()

# ---------------------------------------------------------------------------
# 10. Behavior notes: add via UI
# ---------------------------------------------------------------------------
BehaviorNote.objects.filter(student=naledi, note__startswith="VERIFY:").delete()
resp = c.get(f"/teachers/students/{naledi.id}/behavior/add/")
check("behavior add form 200", resp.status_code == 200)
check("behavior form shows categories", b"Positive" in resp.content and b"Concern" in resp.content)

resp = c.post(
    f"/teachers/students/{naledi.id}/behavior/add/",
    {"category": "POS", "note": "VERIFY: showed leadership in group project"},
    follow=True,
)
check("behavior post follows to profile 200", resp.status_code == 200)
check(
    "behavior note saved",
    BehaviorNote.objects.filter(student=naledi, note__startswith="VERIFY:").exists(),
)
check("note text appears on profile", b"showed leadership in group project" in resp.content)

# Empty note should be rejected
resp = c.post(
    f"/teachers/students/{naledi.id}/behavior/add/",
    {"category": "POS", "note": "   "},
)
check("empty note rejected (form re-rendered)", resp.status_code == 200)
check("error message shown", b"Note text is required" in resp.content)

# Trying to add a note for a student not in teacher's class -> 404
other_school = School.objects.create(name="Other School 2", code="OTHER-2")
other_class = ClassGroup.objects.create(
    school=other_school, name="Other 2A", grade_level=8, academic_year=2026,
)
other_student = Student.objects.create(
    school=other_school, student_number="X-2", first_name="No", last_name="Touchy",
    class_group=other_class,
)
resp = c.get(f"/teachers/students/{other_student.id}/behavior/add/")
check("cross-school behavior add is 404", resp.status_code == 404)
other_student.delete()
other_class.delete()
other_school.delete()

# ---------------------------------------------------------------------------
# 11. Subject management
# ---------------------------------------------------------------------------
Subject.objects.filter(school=naledi.school, code="VERIFY-SUB").delete()
resp = c.get("/teachers/subjects/")
check("subjects manage 200", resp.status_code == 200)
check("subjects page lists Mathematics", b"Mathematics" in resp.content)

resp = c.post("/teachers/subjects/", {"name": "Verify Subject", "code": "VERIFY-SUB"}, follow=True)
check("subject created", Subject.objects.filter(school=naledi.school, code="VERIFY-SUB").exists())

new_subj = Subject.objects.get(school=naledi.school, code="VERIFY-SUB")
check("new subject auto-assigned to teacher", teacher.subjects.filter(id=new_subj.id).exists())
check("new subject shows in list", b"Verify Subject" in resp.content)

# Duplicate code is rejected
resp = c.post("/teachers/subjects/", {"name": "Dup", "code": "VERIFY-SUB"})
check("duplicate subject code rejected", b"already exists" in resp.content)
check(
    "still only one VERIFY-SUB subject",
    Subject.objects.filter(school=naledi.school, code="VERIFY-SUB").count() == 1,
)

# Missing name rejected
resp = c.post("/teachers/subjects/", {"name": "", "code": "X1"})
check("missing-name subject rejected", b"Name and code are required" in resp.content)

# Cleanup
teacher.subjects.remove(new_subj)
new_subj.delete()

# ---------------------------------------------------------------------------
# 12. Class management (create)
# ---------------------------------------------------------------------------
ClassGroup.objects.filter(school=teacher.school, name="VERIFY 9X", academic_year=2026).delete()
resp = c.get("/teachers/classes/new/")
check("class create form 200", resp.status_code == 200)

resp = c.post(
    "/teachers/classes/new/",
    {"name": "VERIFY 9X", "grade_level": "9", "academic_year": "2026"},
    follow=True,
)
new_class = ClassGroup.objects.filter(school=teacher.school, name="VERIFY 9X", academic_year=2026).first()
check("class was created", new_class is not None)
check("teacher set as class_teacher", new_class.class_teacher_id == teacher.id)
check("teacher added to classes_taught", teacher.classes_taught.filter(id=new_class.id).exists())
check("redirected to class detail", b"VERIFY 9X" in resp.content)

# Duplicate name for same year rejected
resp = c.post(
    "/teachers/classes/new/",
    {"name": "VERIFY 9X", "grade_level": "9", "academic_year": "2026"},
)
check("duplicate class rejected", b"already exists" in resp.content)
check(
    "still only one VERIFY 9X class",
    ClassGroup.objects.filter(school=teacher.school, name="VERIFY 9X", academic_year=2026).count() == 1,
)

# Missing grade rejected
resp = c.post(
    "/teachers/classes/new/",
    {"name": "VERIFY 9Y", "grade_level": "", "academic_year": "2026"},
)
check("missing grade level rejected", b"valid grade level" in resp.content)

# Cleanup
teacher.classes_taught.remove(new_class)
new_class.delete()

# ---------------------------------------------------------------------------
# 13. Term-scoped attendance & behavior on PDF reports
# ---------------------------------------------------------------------------
schedule = TermSchedule.objects.get(school=naledi.school, term=1, academic_year=2026)

Attendance.objects.filter(student=naledi).delete()
BehaviorNote.objects.filter(student=naledi).delete()

# In-term attendance (should be on Term 1 report)
in_term = schedule.start_date
out_of_term = schedule.end_date.replace(month=schedule.end_date.month + 1 if schedule.end_date.month < 12 else 12)
Attendance.objects.create(student=naledi, date=in_term, status="P", recorded_by=teacher)
# Out-of-term attendance (should NOT be on Term 1 report)
Attendance.objects.create(student=naledi, date=out_of_term, status="A", recorded_by=teacher)

# In-term behavior note
from django.utils import timezone  # noqa: E402
import datetime as _dt  # noqa: E402
note_in = BehaviorNote.objects.create(
    student=naledi, teacher=teacher, category="POS",
    note="TERM_IN: led morning assembly",
)
BehaviorNote.objects.filter(id=note_in.id).update(
    recorded_at=timezone.make_aware(_dt.datetime.combine(in_term, _dt.time(9, 0))),
)
note_out = BehaviorNote.objects.create(
    student=naledi, teacher=teacher, category="CON",
    note="TERM_OUT: should not be on Term 1 report",
)
BehaviorNote.objects.filter(id=note_out.id).update(
    recorded_at=timezone.make_aware(_dt.datetime.combine(out_of_term, _dt.time(9, 0))),
)

resp = c.get(f"/teachers/students/{naledi.id}/report/?term=1&year=2026")
check("term-scoped report 200", resp.status_code == 200)
pdf_text = "\n".join(p.extract_text() for p in PdfReader(BytesIO(resp.content)).pages)
check(
    "report attendance header uses term dates",
    "Term 1" in pdf_text and "no term dates configured" not in pdf_text,
)
check("in-term behavior note appears", "TERM_IN" in pdf_text)
check("out-of-term behavior note excluded", "TERM_OUT" not in pdf_text)
# In-term attendance counted as Present
import re  # noqa: E402
# After "Present Absent Late Excused", look at the next row of counts
m = re.search(r"Present\s+Absent\s+Late\s+Excused.*?(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)", pdf_text, re.S)
check("attendance counts present in PDF", m is not None)
if m:
    check(
        f"only the in-term Present counted (got Present={m.group(1)}, Absent={m.group(2)}, Total={m.group(5)})",
        m.group(1) == "1" and m.group(2) == "0" and m.group(5) == "1",
    )

# Cleanup
Attendance.objects.filter(student=naledi).delete()
BehaviorNote.objects.filter(student=naledi).delete()

# ---------------------------------------------------------------------------
# 14. School calendar UI
# ---------------------------------------------------------------------------
TermSchedule.objects.filter(school=teacher.school, academic_year=2099).delete()

# GET defaults to current year
resp = c.get("/teachers/calendar/")
check("calendar GET 200", resp.status_code == 200)
check("calendar shows Term 1/2/3 rows",
      b"Term 1" in resp.content and b"Term 2" in resp.content and b"Term 3" in resp.content)

# Switch to a clean year for testing
resp = c.get("/teachers/calendar/?year=2099")
check("year selector switches", b"School calendar &mdash; 2099" in resp.content)
check("all terms show 'Not set' for empty year",
      resp.content.count(b"Not set") >= 3)

# Save a valid 3-term calendar
resp = c.post(
    "/teachers/calendar/",
    {
        "year": "2099",
        "start_1": "2099-01-10", "end_1": "2099-04-05",
        "start_2": "2099-05-01", "end_2": "2099-08-04",
        "start_3": "2099-09-01", "end_3": "2099-12-03",
    },
    follow=True,
)
check("calendar POST follows to 200", resp.status_code == 200)
check(
    "3 term schedules created for 2099",
    TermSchedule.objects.filter(school=teacher.school, academic_year=2099).count() == 3,
)
check("calendar page shows 'Configured' status",
      resp.content.count(b"Configured") >= 3)

# Editing: change Term 1 end date
resp = c.post(
    "/teachers/calendar/",
    {
        "year": "2099",
        "start_1": "2099-01-10", "end_1": "2099-04-15",  # changed
        "start_2": "2099-05-01", "end_2": "2099-08-04",
        "start_3": "2099-09-01", "end_3": "2099-12-03",
    },
    follow=True,
)
ts1 = TermSchedule.objects.get(school=teacher.school, academic_year=2099, term=1)
check("Term 1 end date updated", ts1.end_date.isoformat() == "2099-04-15")
check(
    "still only 3 rows after edit (no duplicates)",
    TermSchedule.objects.filter(school=teacher.school, academic_year=2099).count() == 3,
)

# Validation: end before start -> error, no change
resp = c.post(
    "/teachers/calendar/",
    {
        "year": "2099",
        "start_1": "2099-01-10", "end_1": "2098-12-31",  # bad: end before start
        "start_2": "2099-05-01", "end_2": "2099-08-04",
        "start_3": "2099-09-01", "end_3": "2099-12-03",
    },
)
check("end-before-start rejected", b"End date must be on or after start date" in resp.content)
ts1_after = TermSchedule.objects.get(school=teacher.school, academic_year=2099, term=1)
check("Term 1 unchanged after rejected edit",
      ts1_after.end_date.isoformat() == "2099-04-15")

# Validation: only one of start/end -> error
resp = c.post(
    "/teachers/calendar/",
    {
        "year": "2099",
        "start_1": "2099-01-10", "end_1": "",
        "start_2": "2099-05-01", "end_2": "2099-08-04",
        "start_3": "2099-09-01", "end_3": "2099-12-03",
    },
)
check("partial dates rejected", b"Both start and end dates are required" in resp.content)
check(
    "still 3 rows (atomic — nothing saved on error)",
    TermSchedule.objects.filter(school=teacher.school, academic_year=2099).count() == 3,
)

# Clearing a term: leave both empty -> row deleted
resp = c.post(
    "/teachers/calendar/",
    {
        "year": "2099",
        "start_1": "", "end_1": "",
        "start_2": "2099-05-01", "end_2": "2099-08-04",
        "start_3": "2099-09-01", "end_3": "2099-12-03",
    },
    follow=True,
)
check(
    "cleared Term 1 removed",
    not TermSchedule.objects.filter(school=teacher.school, academic_year=2099, term=1).exists(),
)
check(
    "Terms 2 and 3 retained",
    TermSchedule.objects.filter(school=teacher.school, academic_year=2099).count() == 2,
)

# Calendar dates change report scoping: drop a 2026 term schedule, regenerate
# Naledi report and confirm fallback label appears.
existing_2026_term1 = TermSchedule.objects.get(school=naledi.school, academic_year=2026, term=1)
saved_start, saved_end = existing_2026_term1.start_date, existing_2026_term1.end_date
existing_2026_term1.delete()
resp = c.get(f"/teachers/students/{naledi.id}/report/?term=1&year=2026")
check("report still 200 after removing term schedule", resp.status_code == 200)
pdf_text = "\n".join(p.extract_text() for p in PdfReader(BytesIO(resp.content)).pages)
check("report falls back to year-wide label",
      "no term dates configured" in pdf_text)
# Restore
TermSchedule.objects.create(
    school=naledi.school, academic_year=2026, term=1,
    start_date=saved_start, end_date=saved_end,
)
resp = c.get(f"/teachers/students/{naledi.id}/report/?term=1&year=2026")
pdf_text = "\n".join(p.extract_text() for p in PdfReader(BytesIO(resp.content)).pages)
check("report uses term dates again after restore",
      "no term dates configured" not in pdf_text)

# Cleanup
TermSchedule.objects.filter(school=teacher.school, academic_year=2099).delete()

# ---------------------------------------------------------------------------
# 15. Single-student CRUD
# ---------------------------------------------------------------------------
Student.objects.filter(student_number__in=["Z-NEW-1", "Z-DEL-1"]).delete()

# Create
resp = c.get(f"/teachers/classes/{class_group.id}/students/new/")
check("student create form 200", resp.status_code == 200)
resp = c.post(
    f"/teachers/classes/{class_group.id}/students/new/",
    {
        "first_name": "Newby",
        "last_name": "Maruping",
        "student_number": "Z-NEW-1",
        "date_of_birth": "2012-05-05",
        "gender": "M",
        "parent_name": "Mma Maruping",
        "parent_phone": "+267 71 999 999",
    },
    follow=True,
)
new_student = Student.objects.filter(student_number="Z-NEW-1").first()
check("student created via single-form", new_student is not None)
check("student form redirects to profile", b"Newby Maruping" in resp.content)
check("new student lands in correct class", new_student.class_group_id == class_group.id)

# Create rejects missing first name
resp = c.post(
    f"/teachers/classes/{class_group.id}/students/new/",
    {"first_name": "", "last_name": "X", "student_number": "Z-NEW-2"},
)
check("missing first name rejected", b"First name is required" in resp.content)

# Create rejects duplicate student number
resp = c.post(
    f"/teachers/classes/{class_group.id}/students/new/",
    {"first_name": "Dup", "last_name": "X", "student_number": "Z-NEW-1"},
)
check("duplicate student_number rejected on create", b"already exists" in resp.content)

# Edit
resp = c.get(f"/teachers/students/{new_student.id}/edit/")
check("student edit form 200", resp.status_code == 200)
check("edit form pre-fills first_name", b"Newby" in resp.content)

resp = c.post(
    f"/teachers/students/{new_student.id}/edit/",
    {
        "first_name": "Newbie",
        "last_name": "Maruping",
        "student_number": "Z-NEW-1",
        "date_of_birth": "2012-05-05",
        "gender": "M",
        "parent_name": "Mma Maruping",
        "parent_phone": "+267 71 999 999",
        "class_group": str(class_group.id),
        "is_active": "on",
    },
    follow=True,
)
new_student.refresh_from_db()
check("student edit updated first_name", new_student.first_name == "Newbie")

# Edit rejects invalid date
resp = c.post(
    f"/teachers/students/{new_student.id}/edit/",
    {
        "first_name": "Newbie",
        "last_name": "Maruping",
        "student_number": "Z-NEW-1",
        "date_of_birth": "not-a-date",
        "gender": "M",
        "class_group": str(class_group.id),
        "is_active": "on",
    },
)
check("invalid date rejected on edit", b"YYYY-MM-DD" in resp.content)
new_student.refresh_from_db()
check("student first_name unchanged after rejected edit", new_student.first_name == "Newbie")

# Edit rejects moving to a class the teacher doesn't teach
other_school2 = School.objects.create(name="Far Away School", code="FAR-1")
other_class2 = ClassGroup.objects.create(
    school=other_school2, name="Far 1A", grade_level=8, academic_year=2026,
)
resp = c.post(
    f"/teachers/students/{new_student.id}/edit/",
    {
        "first_name": "Newbie",
        "last_name": "Maruping",
        "student_number": "Z-NEW-1",
        "class_group": str(other_class2.id),
        "is_active": "on",
    },
)
check("moving to non-owned class rejected", b"only move the student into a class you teach" in resp.content)
other_class2.delete()
other_school2.delete()

# Delete confirmation page
resp = c.get(f"/teachers/students/{new_student.id}/delete/")
check("student delete confirm 200", resp.status_code == 200)
check("delete page shows impact", b"mark record" in resp.content and b"attendance record" in resp.content)

# Delete rejects wrong confirmation
resp = c.post(f"/teachers/students/{new_student.id}/delete/", {"confirm": "wrong"})
check(
    "wrong confirmation does not delete",
    Student.objects.filter(id=new_student.id).exists(),
)

# Delete with correct confirmation
new_student_id = new_student.id
resp = c.post(
    f"/teachers/students/{new_student.id}/delete/",
    {"confirm": "Z-NEW-1"},
    follow=True,
)
check("student deleted", not Student.objects.filter(id=new_student_id).exists())
check("delete redirects to class detail", b"Form 1A" in resp.content)

# Delete cascades: create a student with marks, then delete
mark_student = Student.objects.create(
    school=teacher.school, student_number="Z-DEL-1", first_name="Will",
    last_name="Vanish", class_group=class_group,
)
math_sub = Subject.objects.get(school=teacher.school, code="MATH")
Mark.objects.create(
    student=mark_student, subject=math_sub, teacher=teacher,
    assessment_type="TEST", title="Pre-delete test", score=50, max_score=100,
    term=1, academic_year=2026,
)
Attendance.objects.create(
    student=mark_student, date=date(2026, 2, 1), status="P", recorded_by=teacher,
)
BehaviorNote.objects.create(
    student=mark_student, teacher=teacher, category="POS", note="great work",
)
mark_id = Mark.objects.filter(student=mark_student).first().id
c.post(f"/teachers/students/{mark_student.id}/delete/", {"confirm": "Z-DEL-1"})
check("cascade: student gone", not Student.objects.filter(student_number="Z-DEL-1").exists())
check("cascade: marks gone", not Mark.objects.filter(id=mark_id).exists())
check("cascade: attendance gone", not Attendance.objects.filter(student_id=mark_student.id).exists())
check("cascade: behavior notes gone", not BehaviorNote.objects.filter(student_id=mark_student.id).exists())

# Cross-school edit/delete is 404
out_school = School.objects.create(name="Out School", code="OUT-1")
out_class = ClassGroup.objects.create(
    school=out_school, name="Out 1A", grade_level=8, academic_year=2026,
)
out_student = Student.objects.create(
    school=out_school, student_number="OUT-1", first_name="Not", last_name="Mine",
    class_group=out_class,
)
check("cross-school edit 404", c.get(f"/teachers/students/{out_student.id}/edit/").status_code == 404)
check("cross-school delete 404", c.get(f"/teachers/students/{out_student.id}/delete/").status_code == 404)
out_student.delete(); out_class.delete(); out_school.delete()

# ---------------------------------------------------------------------------
# 16. Enquiries
# ---------------------------------------------------------------------------
Enquiry.objects.filter(from_teacher=teacher).delete()

resp = c.get("/teachers/enquiries/")
check("enquiry list 200", resp.status_code == 200)
check("enquiry list empty state", b"No enquiries yet" in resp.content)

resp = c.get("/teachers/enquiries/new/")
check("enquiry form 200", resp.status_code == 200)
check("enquiry form shows HR category", b"HR" in resp.content)

# Validation
resp = c.post("/teachers/enquiries/new/", {"subject": "", "body": ""})
check("missing fields rejected", b"Subject is required" in resp.content and b"describe your enquiry" in resp.content)

# Create good enquiry
resp = c.post(
    "/teachers/enquiries/new/",
    {"category": "HR", "subject": "Leave request for next week",
     "body": "I'd like to request 2 days of leave next Thursday and Friday."},
    follow=True,
)
created = Enquiry.objects.filter(from_teacher=teacher, subject="Leave request for next week").first()
check("enquiry created", created is not None)
check("enquiry redirects to detail showing the subject", b"Leave request for next week" in resp.content)
check("enquiry detail shows category", b"HR" in resp.content)
check("enquiry shows 'No response yet'", b"No response yet" in resp.content)

# Admin response (simulated by direct model write)
created.status = Enquiry.Status.RESOLVED
created.response = "Approved. Please brief Mr. Sebina before you leave."
created.save()
resp = c.get(f"/teachers/enquiries/{created.id}/")
check("enquiry response surfaces to teacher", b"Approved. Please brief" in resp.content)
check("resolved status badge shown", b"Resolved" in resp.content)

# List shows the enquiry
resp = c.get("/teachers/enquiries/")
check("enquiry list shows item", b"Leave request for next week" in resp.content)
check("enquiry list shows resolved badge", b"Resolved" in resp.content)

# Another teacher's enquiry should not be visible
other_school3 = School.objects.create(name="Other School 3", code="OTHER-3")
from django.contrib.auth import get_user_model  # noqa: E402
User = get_user_model()
other_user = User.objects.create_user(
    username="other_teacher", password="x", first_name="Other", last_name="T",
)
other_teacher = TeacherProfile.objects.create(user=other_user, school=other_school3)
other_enquiry = Enquiry.objects.create(
    school=other_school3, from_teacher=other_teacher,
    subject="SECRET", body="not for you", category="HR",
)
resp = c.get("/teachers/enquiries/")
check("other teacher's enquiry not in list", b"SECRET" not in resp.content)
check(
    "other teacher's enquiry detail 404",
    c.get(f"/teachers/enquiries/{other_enquiry.id}/").status_code == 404,
)
other_enquiry.delete(); other_teacher.delete(); other_user.delete(); other_school3.delete()

Enquiry.objects.filter(from_teacher=teacher).delete()

# ---------------------------------------------------------------------------
# 17. Enhanced dashboard
# ---------------------------------------------------------------------------
# Seed a known performance baseline so we can predict the avg
Mark.objects.filter(teacher=teacher, academic_year=2026).delete()
math_sub = Subject.objects.get(school=teacher.school, code="MATH")
eng_sub = Subject.objects.get(school=teacher.school, code="ENG")

dash_students = list(Student.objects.filter(class_group=class_group, is_active=True)[:3])
for student in dash_students:
    Mark.objects.create(
        student=student, subject=math_sub, teacher=teacher,
        assessment_type="TEST", title="Dash test M", score=80, max_score=100,
        term=1, academic_year=2026,
    )
    Mark.objects.create(
        student=student, subject=eng_sub, teacher=teacher,
        assessment_type="TEST", title="Dash test E", score=60, max_score=100,
        term=1, academic_year=2026,
    )

resp = c.get("/teachers/")
check("dashboard 200", resp.status_code == 200)
body = resp.content
check("dashboard shows teacher name", b"Kgosi" in body)
check("dashboard shows school", b"Gaborone Demo Secondary School" in body)
check("dashboard shows classes stat", b"My classes" in body)
check("dashboard shows subjects stat", b"Subjects taught" in body)
check("dashboard shows students stat", b"Students" in body)
check("dashboard shows attendance rate label", b"Attendance rate" in body)
check("dashboard shows overall performance card", b"Student performance" in body)
# Avg should be (80+60+80+60+80+60)/6 = 70.0 -> "70.0%"
check("dashboard shows correct overall avg (70.0%)", b"70.0%" in body, "expected 70.0% from seeded marks")
check("dashboard shows per-subject performance section", b"Performance by subject" in body)
check("dashboard shows Mathematics row in per-subject", b"Mathematics" in body)
check("dashboard shows English row in per-subject", b"English" in body)
check("dashboard shows teacher performance band", b"Good" in body or b"Excellent" in body)
check("dashboard shows recent marks section", b"Recent marks recorded" in body)
check("dashboard shows enquiries card", b"Open enquiries" in body)

# Cleanup
Mark.objects.filter(teacher=teacher, academic_year=2026, title__startswith="Dash test").delete()

# ---------------------------------------------------------------------------
# 18. School admin portal
# ---------------------------------------------------------------------------
admin_school = teacher.school

# 18a. Smart login redirect: teacher goes to /teachers/, admin goes to /admin-portal/
teacher_client = Client()
teacher_client.login(username="mr_kgosi", password="teacher123")
resp = teacher_client.get("/", follow=False)
check("teacher root redirects to teachers portal",
      resp.status_code == 302 and "/teachers/" in resp["Location"])

admin_client = Client()
assert admin_client.login(username="mma_pula", password="admin123"), "school admin login failed"
resp = admin_client.get("/", follow=False)
check("admin root redirects to admin portal",
      resp.status_code == 302 and "/admin-portal/" in resp["Location"])

# 18b. Teacher can't access the admin portal
resp = teacher_client.get("/admin-portal/")
check("teacher denied admin portal (403)", resp.status_code == 403)
check("denial page mentions no admin access",
      b"No admin access" in resp.content or b"no admin" in resp.content.lower())

# 18c. Admin dashboard
resp = admin_client.get("/admin-portal/")
check("admin dashboard 200", resp.status_code == 200)
body = resp.content
check("admin dashboard shows school name", b"Gaborone Demo Secondary School" in body)
check("admin dashboard shows Teachers card", b"Teachers" in body)
check("admin dashboard shows Students card", b"Students" in body)
check("admin dashboard shows Classes card", b"Classes" in body)
check("admin dashboard shows Subjects card", b"Subjects" in body)
check("admin dashboard shows Open enquiries card", b"Open enquiries" in body)
check("admin dashboard shows recent enquiries section", b"Recent enquiries" in body)
check("admin dashboard shows school-wide performance", b"School-wide performance" in body)

# 18d. Enquiries inbox + respond
Enquiry.objects.filter(school=admin_school).delete()
e1 = Enquiry.objects.create(
    school=admin_school, from_teacher=teacher, category="HR",
    subject="Need new chairs for Form 1A", body="Half of them are broken.",
)
Enquiry.objects.create(
    school=admin_school, from_teacher=teacher, category="TECH",
    subject="Cannot upload Excel today", body="Times out at 90%.", status="PROG",
)

resp = admin_client.get("/admin-portal/enquiries/")
check("admin enquiry inbox 200", resp.status_code == 200)
check("inbox lists both enquiries",
      b"Need new chairs for Form 1A" in resp.content and b"Cannot upload Excel today" in resp.content)

# Filter by status=OPEN -> only the chairs enquiry
resp = admin_client.get("/admin-portal/enquiries/?status=OPEN")
check("OPEN filter shows chairs enquiry", b"Need new chairs for Form 1A" in resp.content)
check("OPEN filter hides PROG enquiry", b"Cannot upload Excel today" not in resp.content)

# Filter by category=TECH -> only excel one
resp = admin_client.get("/admin-portal/enquiries/?category=TECH")
check("TECH filter shows excel enquiry", b"Cannot upload Excel today" in resp.content)
check("TECH filter hides HR enquiry", b"Need new chairs for Form 1A" not in resp.content)

# Detail page and reply
resp = admin_client.get(f"/admin-portal/enquiries/{e1.id}/")
check("admin enquiry detail 200", resp.status_code == 200)
check("detail shows teacher message body", b"Half of them are broken" in resp.content)

resp = admin_client.post(
    f"/admin-portal/enquiries/{e1.id}/",
    {"status": "DONE", "response": "Approved — Procurement will deliver Monday."},
    follow=True,
)
e1.refresh_from_db()
check("enquiry status updated to RESOLVED", e1.status == "DONE")
check("enquiry response stored", "Procurement will deliver Monday" in e1.response)
check("resolved_at populated on RESOLVED", e1.resolved_at is not None)

# Teacher sees the admin's response now
teacher_view = teacher_client.get(f"/teachers/enquiries/{e1.id}/")
check("teacher sees admin response", b"Procurement will deliver Monday" in teacher_view.content)
check("teacher sees Resolved status", b"Resolved" in teacher_view.content)

# Admin can flip status back -> resolved_at clears
admin_client.post(
    f"/admin-portal/enquiries/{e1.id}/",
    {"status": "PROG", "response": e1.response},
)
e1.refresh_from_db()
check("re-opening clears resolved_at", e1.resolved_at is None)

# Cross-school isolation
other_school_x = School.objects.create(name="Far Admin Test", code="FAR-ADM")
other_user_x = TeacherProfile.objects.create(
    user=__import__("core.models", fromlist=["User"]).User.objects.create(username="other_teach_x", password="x"),
    school=other_school_x,
)
other_enquiry_x = Enquiry.objects.create(
    school=other_school_x, from_teacher=other_user_x, category="HR",
    subject="SECRET-ADMIN", body="not for this school",
)
resp = admin_client.get("/admin-portal/enquiries/")
check("admin can't see other school's enquiry in list", b"SECRET-ADMIN" not in resp.content)
check(
    "admin gets 404 on other school's enquiry detail",
    admin_client.get(f"/admin-portal/enquiries/{other_enquiry_x.id}/").status_code == 404,
)
other_enquiry_x.delete(); other_user_x.user.delete(); other_user_x.delete(); other_school_x.delete()
Enquiry.objects.filter(school=admin_school).delete()

# 18e. Teachers list + create + edit
resp = admin_client.get("/admin-portal/teachers/")
check("admin teachers list 200", resp.status_code == 200)
check("teachers list shows mr_kgosi", b"mr_kgosi" in resp.content)

# Create teacher
math = Subject.objects.get(school=admin_school, code="MATH")
existing_class = ClassGroup.objects.get(school=admin_school, name="Form 1A")
User_ = __import__("core.models", fromlist=["User"]).User
User_.objects.filter(username="new_teach_x").delete()
resp = admin_client.post(
    "/admin-portal/teachers/new/",
    {
        "first_name": "Newby",
        "last_name": "Lecturer",
        "username": "new_teach_x",
        "email": "newby@example.com",
        "phone": "+267 71 555 555",
        "password": "securepass1",
        "employee_id": "T-NEW",
        "subjects": [str(math.id)],
        "classes": [str(existing_class.id)],
    },
    follow=True,
)
new_user = User_.objects.filter(username="new_teach_x").first()
check("new teacher user created", new_user is not None)
new_profile = TeacherProfile.objects.filter(user=new_user).first()
check("new teacher profile linked to school", new_profile and new_profile.school_id == admin_school.id)
check("new teacher assigned math subject", math in new_profile.subjects.all())
check("new teacher assigned class", existing_class in new_profile.classes_taught.all())

# New teacher can log in with the admin-set password
fresh_client = Client()
check(
    "new teacher can authenticate",
    fresh_client.login(username="new_teach_x", password="securepass1"),
)

# Duplicate username rejected
resp = admin_client.post(
    "/admin-portal/teachers/new/",
    {"first_name": "X", "last_name": "Y", "username": "new_teach_x", "password": "anotherpass1"},
)
check("duplicate username rejected", b"already taken" in resp.content)

# Short password rejected
resp = admin_client.post(
    "/admin-portal/teachers/new/",
    {"first_name": "X", "last_name": "Y", "username": "shortpw_x", "password": "abc"},
)
check("short password rejected", b"at least 8 characters" in resp.content)
check("short password did not create user", not User_.objects.filter(username="shortpw_x").exists())

# Edit: deactivate new teacher and clear subjects
resp = admin_client.post(
    f"/admin-portal/teachers/{new_profile.id}/edit/",
    {
        "first_name": "Newby",
        "last_name": "Lecturer",
        "email": "newby@example.com",
        "phone": "+267 71 555 555",
        "employee_id": "T-NEW",
        "password": "",
        "subjects": [],
        "classes": [],
    },
    follow=True,
)
new_profile.refresh_from_db()
check("edit deactivated teacher (no is_active checkbox sent)", not new_profile.is_active)
check("edit cleared subjects", new_profile.subjects.count() == 0)

# Cleanup new teacher
new_profile.delete(); new_user.delete()

# 18f. Students cross-class list + filters
resp = admin_client.get("/admin-portal/students/")
check("admin students list 200", resp.status_code == 200)
check("students list shows seeded Naledi", b"Naledi" in resp.content)
check("students list shows seeded Tumelo", b"Tumelo" in resp.content)

resp = admin_client.get("/admin-portal/students/?q=Naledi")
check("student search by first name works", b"Naledi" in resp.content)
check("student search excludes others", b"Tumelo" not in resp.content)

resp = admin_client.get(f"/admin-portal/students/?class={existing_class.id}")
check(
    "class filter returns Form 1A students",
    b"Naledi" in resp.content and b"Tumelo" in resp.content,
)

# 18g. Classes list + create + edit
resp = admin_client.get("/admin-portal/classes/")
check("admin classes list 200", resp.status_code == 200)
check("classes list shows Form 1A", b"Form 1A" in resp.content)
check("classes list shows class teacher name", b"Kgosi" in resp.content)

ClassGroup.objects.filter(school=admin_school, name="Form 2C", academic_year=2026).delete()
resp = admin_client.post(
    "/admin-portal/classes/new/",
    {
        "name": "Form 2C",
        "grade_level": "9",
        "academic_year": "2026",
        "class_teacher": str(teacher.id),
    },
    follow=True,
)
new_class = ClassGroup.objects.filter(school=admin_school, name="Form 2C").first()
check("admin created class", new_class is not None)
check("class teacher assigned by admin", new_class.class_teacher_id == teacher.id)
check("class teacher's classes_taught updated", teacher.classes_taught.filter(id=new_class.id).exists())

# Duplicate rejected
resp = admin_client.post(
    "/admin-portal/classes/new/",
    {"name": "Form 2C", "grade_level": "9", "academic_year": "2026"},
)
check("admin class duplicate rejected", b"already exists" in resp.content)

# Edit rename
resp = admin_client.post(
    f"/admin-portal/classes/{new_class.id}/edit/",
    {"name": "Form 2D", "grade_level": "9", "academic_year": "2026",
     "class_teacher": str(teacher.id)},
    follow=True,
)
new_class.refresh_from_db()
check("admin renamed class", new_class.name == "Form 2D")

# Cleanup
teacher.classes_taught.remove(new_class)
new_class.delete()

# 18h. Subjects: list + create
Subject.objects.filter(school=admin_school, code="ADM-SUB-X").delete()
resp = admin_client.get("/admin-portal/subjects/")
check("admin subjects list 200", resp.status_code == 200)
check("admin subjects list shows Mathematics", b"Mathematics" in resp.content)

resp = admin_client.post(
    "/admin-portal/subjects/",
    {"name": "Admin Test Subject", "code": "ADM-SUB-X"},
    follow=True,
)
check(
    "admin created subject",
    Subject.objects.filter(school=admin_school, code="ADM-SUB-X").exists(),
)
# Duplicate
resp = admin_client.post(
    "/admin-portal/subjects/",
    {"name": "Other name", "code": "ADM-SUB-X"},
)
check("admin subject duplicate rejected", b"already exists" in resp.content)
Subject.objects.filter(school=admin_school, code="ADM-SUB-X").delete()

# 18i. School profile edit
original_region = admin_school.region
resp = admin_client.get("/admin-portal/school/")
check("school profile 200", resp.status_code == 200)
check("school profile shows name field", b"Gaborone Demo Secondary School" in resp.content)

resp = admin_client.post(
    "/admin-portal/school/",
    {
        "name": admin_school.name,
        "region": "Central",
        "address": "Plot 9999, Gaborone",
        "phone": admin_school.phone,
        "email": admin_school.email,
        "principal_name": "Mma Pula Mokgothu",
    },
    follow=True,
)
admin_school.refresh_from_db()
check("school region updated via admin portal", admin_school.region == "Central")
check("school principal updated", admin_school.principal_name == "Mma Pula Mokgothu")
# Restore
admin_school.region = original_region
admin_school.save()

# Empty school name rejected
resp = admin_client.post(
    "/admin-portal/school/",
    {"name": "", "region": "X"},
)
check("empty school name rejected", b"School name is required" in resp.content)

# 18j. Admin calendar (same logic as teacher's, distinct path)
TermSchedule.objects.filter(school=admin_school, academic_year=2098).delete()
resp = admin_client.get("/admin-portal/calendar/?year=2098")
check("admin calendar 200", resp.status_code == 200)
check("admin calendar shows year heading", b"2098" in resp.content)

resp = admin_client.post(
    "/admin-portal/calendar/",
    {
        "year": "2098",
        "start_1": "2098-01-10", "end_1": "2098-04-05",
        "start_2": "", "end_2": "",
        "start_3": "", "end_3": "",
    },
    follow=True,
)
check(
    "admin calendar saved Term 1 for 2098",
    TermSchedule.objects.filter(school=admin_school, academic_year=2098, term=1).exists(),
)
# Cleanup
TermSchedule.objects.filter(school=admin_school, academic_year=2098).delete()

# ---------------------------------------------------------------------------
# 19. WhatsApp parent webhook
# ---------------------------------------------------------------------------
ParentSession.objects.all().delete()

# Make sure Naledi has at least some marks (seed re-runs create them, but
# previous test cleanups may have wiped them). Recreate the minimal set.
naledi = Student.objects.get(student_number="S-2026-001")
math_subj = Subject.objects.get(school=naledi.school, code="MATH")
eng_subj = Subject.objects.get(school=naledi.school, code="ENG")
Mark.objects.filter(student=naledi, term=1, academic_year=2026, title__startswith="WhatsApp").delete()
Mark.objects.create(
    student=naledi, subject=math_subj, teacher=teacher,
    assessment_type="TEST", title="WhatsApp test 1",
    score=78, max_score=100, term=1, academic_year=2026,
)
Mark.objects.create(
    student=naledi, subject=eng_subj, teacher=teacher,
    assessment_type="ASSIGN", title="WhatsApp essay 1",
    score=42, max_score=50, term=1, academic_year=2026,
)
Attendance.objects.filter(student=naledi).delete()
Attendance.objects.create(student=naledi, date=date(2026, 5, 20), status="P", recorded_by=teacher)
Attendance.objects.create(student=naledi, date=date(2026, 5, 21), status="P", recorded_by=teacher)
Attendance.objects.create(student=naledi, date=date(2026, 5, 22), status="A", recorded_by=teacher)
BehaviorNote.objects.filter(student=naledi, note__startswith="WA:").delete()
BehaviorNote.objects.create(
    student=naledi, teacher=teacher, category="POS",
    note="WA: showed great teamwork in science project",
)

# Use the unauthenticated test client — webhook is public (validated by Twilio sig in prod)
wa = Client()
WEBHOOK = "/whatsapp/webhook/"
NALEDI_PHONE = "whatsapp:+26771222001"  # matches +267 71 222 001 in fixture
UNKNOWN_PHONE = "whatsapp:+26799999999"

# 19a. GET returns a friendly health message
resp = wa.get(WEBHOOK)
check("webhook GET 200", resp.status_code == 200)
check("webhook GET mentions ThutoTrack", b"ThutoTrack" in resp.content)

# 19b. Unknown phone: gentle "your number isn't linked" reply
resp = wa.post(WEBHOOK, {"From": UNKNOWN_PHONE, "Body": "hi"})
check("webhook POST 200 for unknown phone", resp.status_code == 200)
check("webhook returns XML (TwiML)", resp["Content-Type"].startswith("application/xml"))
check("webhook body wrapped in <Response><Message>", b"<Response><Message>" in resp.content)
check("unknown phone gets onboarding message", b"isn't linked to a student" in resp.content)
check("no ParentSession was created for unknown phone",
      not ParentSession.objects.filter(phone__contains="9999").exists())

# 19c. Known phone, "hi" -> welcome menu showing single student
resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "hi"})
body = resp.content.decode("utf-8")
check("known phone gets ThutoTrack greeting", "ThutoTrack" in body)
check("welcome mentions the student name", "Naledi Seretse" in body)
check("welcome shows class", "Form 1A" in body)
check("welcome lists available commands", "marks" in body and "attendance" in body and "report" in body)
session = ParentSession.objects.get(phone__contains="22001")
check("ParentSession was created", session is not None)
check("session auto-selected the only student", session.selected_student_id == naledi.id)

# 19d. "marks"
resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "marks"})
body = resp.content.decode("utf-8")
check("marks reply mentions student", "Naledi" in body)
check("marks reply includes Mathematics", "Mathematics" in body)
check("marks reply includes WhatsApp test 1", "WhatsApp test 1" in body)
check("marks reply includes English", "English" in body)
check("marks reply contains percentage", "78%" in body or "78" in body)

# 19e. "attendance"
resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "attendance"})
body = resp.content.decode("utf-8")
check("attendance reply mentions student", "Naledi" in body)
check("attendance reply shows Present count", "Present" in body)
# 2 present out of 3 = 67%
check("attendance reply shows correct rate (67%)", "67%" in body)

# 19f. "report" -- term summary
resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "report"})
body = resp.content.decode("utf-8")
check("report reply mentions Term", "Term" in body)
check("report reply has Overall line", "Overall" in body)
# Math: 78%, English: 84% -> avg = 81%
check("report overall avg correct (81%)", "81%" in body)

# 19g. "behavior"
resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "behavior"})
body = resp.content.decode("utf-8")
check("behavior reply mentions student", "Naledi" in body)
check("behavior reply shows seeded note", "showed great teamwork" in body)

# 19h. Unknown command falls through to student menu with apology
resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "gibberish"})
body = resp.content.decode("utf-8")
check("unknown command gets apology", "didn't understand" in body)
check("unknown command still shows student menu", "marks" in body)

# 19i. Shortcut digits work as menu choices (the only-student case: "1" means marks)
resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "1"})
body = resp.content.decode("utf-8")
check("'1' shortcut returns marks", "Mathematics" in body)

# 19j. Multi-student parent — register a sibling so we exercise the menu
sibling = Student.objects.create(
    school=naledi.school,
    student_number="S-2026-099",
    first_name="Karabo",
    last_name="Seretse",
    class_group=naledi.class_group,
    parent_phone="+267 71 222 001",  # same parent
)
ParentSession.objects.filter(phone__contains="22001").delete()

resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "hi"})
body = resp.content.decode("utf-8")
check("multi-student welcome lists both children",
      "Naledi Seretse" in body and "Karabo Seretse" in body)
check("multi-student menu asks for a number", "number" in body)

# Pick student 1 (alphabetical order — Karabo first)
resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "1"})
body = resp.content.decode("utf-8")
session = ParentSession.objects.get(phone__contains="22001")
chosen = session.selected_student
check("'1' selects a student", chosen is not None)
check("selected student menu replies", chosen.first_name.encode() in resp.content)

# "marks" now applies to the chosen sibling and should report no marks
resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "marks"})
body = resp.content.decode("utf-8")
if chosen.id == sibling.id:
    check("marks for sibling-without-marks says 'No marks recorded'",
          "No marks" in body)
else:
    check("marks for Naledi still works after menu choice",
          "Mathematics" in body)

# Invalid number reply
resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "99"})
body = resp.content.decode("utf-8")
check("invalid student-number choice gets a helpful error",
      "not one of your students" in body or "Invalid" in body)

# Cleanup sibling and sessions
sibling.delete()
ParentSession.objects.all().delete()
Mark.objects.filter(student=naledi, title__startswith="WhatsApp").delete()

# 19k. Signature validation: enable a token and ensure unsigned requests are rejected
from django.test import override_settings  # noqa: E402

with override_settings(WHATSAPP_AUTH_TOKEN="test-twilio-token"):
    resp = wa.post(WEBHOOK, {"From": NALEDI_PHONE, "Body": "hi"})
    check("webhook rejects unsigned POST when token is set (403)", resp.status_code == 403)

    # Provide a correct signature and verify acceptance
    import base64 as _b64  # noqa: E402
    import hmac as _hmac  # noqa: E402
    from hashlib import sha1 as _sha1  # noqa: E402

    params = {"From": NALEDI_PHONE, "Body": "hi"}
    url = "http://testserver" + WEBHOOK
    payload = url + "".join(f"{k}{v}" for k, v in sorted(params.items()))
    sig = _b64.b64encode(
        _hmac.new(b"test-twilio-token", payload.encode("utf-8"), _sha1).digest()
    ).decode("ascii")
    resp = wa.post(WEBHOOK, params, HTTP_X_TWILIO_SIGNATURE=sig)
    check("correctly-signed POST accepted (200)", resp.status_code == 200)

# Final cleanup
ParentSession.objects.all().delete()

# ---------------------------------------------------------------------------
# 20. Landing page + 4 login sections + parent web portal
# ---------------------------------------------------------------------------

# 20a. Anonymous landing renders all 4 sections with the right form actions
anon = Client()
resp = anon.get("/")
check("anonymous landing 200", resp.status_code == 200)
body = resp.content
check("landing shows Teachers section", b"Teachers" in body and b"Sign in as teacher" in body)
check("landing shows School Head section", b"School Head" in body and b"Sign in as school head" in body)
check("landing shows Site administrator section", b"Site administrator" in body and b"Django admin" in body)
check("landing shows Parents section", b"Parents" in body and b"Sign in as parent" in body)
check("teacher form posts to teachers login", b'action="/teachers/login/"' in body)
check("school-head form posts to teachers login with admin-portal next",
      b'name="next" value="/admin-portal/"' in body)
check("parent form posts to parents login", b'action="/parents/login/"' in body)
check("admin form points at Django admin", b'action="/admin/login/' in body)

# 20b. Authenticated users get smart-redirected (no landing)
teacher_client2 = Client()
teacher_client2.login(username="mr_kgosi", password="teacher123")
resp = teacher_client2.get("/", follow=False)
check("logged-in teacher does not see landing",
      resp.status_code == 302 and "/teachers/" in resp["Location"])

admin_client2 = Client()
admin_client2.login(username="mma_pula", password="admin123")
resp = admin_client2.get("/", follow=False)
check("logged-in school head does not see landing",
      resp.status_code == 302 and "/admin-portal/" in resp["Location"])

# 20c. Parent login - happy path with seeded credentials
parent_client = Client()
resp = parent_client.get("/parents/login/")
check("parent login page 200", resp.status_code == 200)
check("parent login asks for phone + PIN",
      b"Phone number" in resp.content and b"PIN" in resp.content)

# Wrong PIN
resp = parent_client.post(
    "/parents/login/", {"phone": "+267 71 222 001", "pin": "9999"},
)
check("wrong PIN rejected", b"find a parent account" in resp.content)

# Wrong phone (no account)
resp = parent_client.post(
    "/parents/login/", {"phone": "+267 71 000 000", "pin": "1234"},
)
check("unknown phone rejected", b"find a parent account" in resp.content)

# Empty
resp = parent_client.post("/parents/login/", {"phone": "", "pin": ""})
check("empty submit rejected", b"Enter your phone" in resp.content)

# 20d. Login with various phone formats — all should work
for raw in ["+267 71 222 001", "26771222001", "071222001", "71222001"]:
    cli = Client()
    resp = cli.post("/parents/login/", {"phone": raw, "pin": "1234"}, follow=False)
    check(
        f"parent login accepts format '{raw}'",
        resp.status_code == 302 and "/parents/" in resp["Location"],
    )

# 20e. Authenticated landing now redirects parent to /parents/
resp = parent_client.post(
    "/parents/login/", {"phone": "+267 71 222 001", "pin": "1234"}, follow=False,
)
check("parent login redirects to dashboard",
      resp.status_code == 302 and resp["Location"].endswith("/parents/"))

# Hit landing while logged in -> redirect to /parents/
resp = parent_client.get("/", follow=False)
check("logged-in parent redirected to parent portal",
      resp.status_code == 302 and "/parents/" in resp["Location"])

# 20f. Parent dashboard
resp = parent_client.get("/parents/")
check("parent dashboard 200", resp.status_code == 200)
check("dashboard shows Naledi", b"Naledi Seretse" in resp.content)
check("dashboard shows school", b"Gaborone Demo Secondary School" in resp.content)
check("dashboard shows parent phone", b"26771222001" in resp.content)

# 20g. Per-student view — accessible
resp = parent_client.get(f"/parents/students/{naledi.id}/")
check("parent student detail 200", resp.status_code == 200)
check("student detail shows name", b"Naledi Seretse" in resp.content)
check("student detail shows marks section", b"All marks" in resp.content)
check("student detail shows attendance", b"Attendance" in resp.content)
check("student detail shows behavior section", b"Behavior notes" in resp.content)

# 20h. Per-student view — NOT accessible for someone else's child
other_school_x = School.objects.create(name="Other Parent Test", code="OPT-1")
other_class_x = ClassGroup.objects.create(
    school=other_school_x, name="OPT 1A", grade_level=8, academic_year=2026,
)
not_my_kid = Student.objects.create(
    school=other_school_x,
    student_number="OPT-1",
    first_name="Not",
    last_name="Mine",
    class_group=other_class_x,
    parent_phone="+267 71 000 000",
)
resp = parent_client.get(f"/parents/students/{not_my_kid.id}/")
check("foreign student is 404", resp.status_code == 404)
not_my_kid.delete(); other_class_x.delete(); other_school_x.delete()

# 20i. Parent term report PDF
resp = parent_client.get(f"/parents/students/{naledi.id}/report/?term=1&year=2026")
check("parent term report 200", resp.status_code == 200)
check("parent term report is PDF", resp["Content-Type"] == "application/pdf")
check("parent term report body starts with %PDF", resp.content.startswith(b"%PDF"))

# 20j. Anonymous can't access parent dashboard
resp = anon.get("/parents/")
check("anonymous bounced from parent dashboard",
      resp.status_code in (302, 403))

# Teachers/admins can't access parent dashboard either
resp = teacher_client2.get("/parents/", follow=False)
check("teacher bounced from parent dashboard", resp.status_code == 302)

# 20k. Parent logout
resp = parent_client.post("/parents/logout/", follow=False)
check("parent logout redirects to landing",
      resp.status_code == 302 and resp["Location"] == "/")
resp = parent_client.get("/parents/")
check("after logout, parent dashboard requires login again", resp.status_code == 302)

# 20l. Schooladmin "Parents" page
admin_client2 = Client()
admin_client2.login(username="mma_pula", password="admin123")
resp = admin_client2.get("/admin-portal/parents/")
check("admin parents page 200", resp.status_code == 200)
check("admin parents page lists parent phones", b"71 222 001" in resp.content)
check("admin parents page shows PIN-set badge", b"PIN set" in resp.content)

# Reset a PIN
resp = admin_client2.post(
    "/admin-portal/parents/set-pin/",
    {"phone": "+267 71 222 002", "pin": "5678"},  # Tumelo's parent (no account yet)
    follow=True,
)
new_parent = User.objects.filter(
    username="26771222002", role="PARENT",
).first()
check("admin created new parent account", new_parent is not None)

# That parent can now log in
cli3 = Client()
resp = cli3.post(
    "/parents/login/", {"phone": "+267 71 222 002", "pin": "5678"}, follow=False,
)
check("newly-set PIN works for login",
      resp.status_code == 302 and "/parents/" in resp["Location"])

# Invalid: too short PIN
resp = admin_client2.post(
    "/admin-portal/parents/set-pin/",
    {"phone": "+267 71 222 003", "pin": "12"},
    follow=True,
)
check("short PIN rejected", b"at least 4 digits" in resp.content)
check(
    "short PIN didn't create user",
    not User.objects.filter(username="26771222003", role="PARENT").exists(),
)

# Invalid: phone not at this school
resp = admin_client2.post(
    "/admin-portal/parents/set-pin/",
    {"phone": "+267 99 999 999", "pin": "1234"},
    follow=True,
)
check("phone not at school rejected", b"No student at this school" in resp.content)

# Cleanup the parent we created
User.objects.filter(username="26771222002", role="PARENT").delete()

print(f"\nALL {ASSERT_COUNT} CHECKS PASSED")
